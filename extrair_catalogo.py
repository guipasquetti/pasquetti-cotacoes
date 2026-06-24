"""
extrair_catalogo.py
Parseia as tabelas de preço Excel (HL Consumo, HL Revenda, Pressão JGS)
e gera catalogo_produtos.json.

Rode pelo "3_ATUALIZAR PREÇOS.command" sempre que receber novas tabelas.
Coloque os arquivos Excel na MESMA pasta deste script. Os nomes podem variar
desde que contenham as palavras CONSUMO / REVENDA / PRESS no nome do arquivo.
"""
import openpyxl
import json
import re
import os
import glob

# Pasta onde este script (e os Excel) estão — funciona em qualquer máquina.
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))


# ════════════════════════════════════════════════════════════════════════════
# CLASSIFICAÇÃO DE CÉLULAS  (cabeçalho de categoria  vs  tamanho/item)
# ════════════════════════════════════════════════════════════════════════════
# Regra central, robusta:
#   • É CABEÇALHO de categoria  → quando NÃO tem preço E contém ao menos uma
#     palavra com 3+ letras (ex.: "Curva BB 90º", "TUBO K7 ÁGUA", "Valv. Gav.").
#   • É TAMANHO/ITEM            → quando tem preço, ou é um código de tamanho
#     puramente dimensional (ex.: "100mm", "80x50", "A - 50", "5/8x3\"").
#
# O bug antigo classificava "Curva BB 90º" e "TUBO K7 ÁGUA" como TAMANHO
# (porque eram curtos e tinham número), jogando dezenas de produtos no rótulo
# genérico "JGS". A nova regra usa a presença de palavra alfabética longa.

# Padrões de tamanho puramente dimensionais (sem palavra de produto).
SIZE_PATTERNS = [
    r'^\d+\s*mm$',                      # 50mm, 100 mm
    r'^\d+\s*mm\s*P/[PB]$',             # 200mm P/P, 100mm P/B
    r'^\d+\s*x\s*\d+\s*mm?$',           # 75 x 50mm, 80x50
    r'^\d+\s*x\s*\d+$',                 # 80x50, 250x250
    r'^[A-K]\s*-\s*\d+$',               # A - 50 ... K - 300 (Ultralink)
    r'^\d+x\d+/\d+x\d+$',               # 50x1/2x3/4 (colar de tomada)
    r'^\d+/\d+\s*x\s*[\d.\s/"]+$',      # 5/8x3", 5/8x3 1/2, 3/4x4.1/2 (parafusos)
    r'^\d+/\d+mm$',                     # 75/80mm
    r'^TDA-\d+.*$',                     # TDA-100 ...
    r'^SG\s*\d+x\d+$',                  # SG 100x50
    r'^\d+\s*(g|kg)$',                  # 500g (lubrificante)
]

def clean_text(val):
    if val is None:
        return None
    return str(val).strip()

def is_strict_size(text):
    """True só para códigos de tamanho dimensionais (sem palavra de produto)."""
    if not text:
        return False
    t = str(text).strip()
    return any(re.match(p, t, re.IGNORECASE) for p in SIZE_PATTERNS)

def is_header(nome, preco):
    """
    True se a célula é um CABEÇALHO de categoria.
    Cabeçalho = sem preço E tem ao menos uma palavra com 3+ letras
    (e não é um código de tamanho dimensional).
    """
    if preco is not None:
        return False
    if is_strict_size(nome):
        return False
    # palavra com 3 ou mais letras (acentos incluídos)
    return bool(re.search(r'[A-Za-zÀ-ÿ]{3,}', nome))


# ════════════════════════════════════════════════════════════════════════════
# PARSERS
# ════════════════════════════════════════════════════════════════════════════

SKIP_TEXTS = {
    'IMPOSTOS INCLUSOS', 'PAGAMENTO 30 DIAS',
    'ENTREGA SUJEITA A CONFIRMAÇÃO DE ESTOQUE',
    'PREÇOS SUJEITOS A ALTERAÇÃO SEM AVISO PRÉVIO', 'TUBOS   18% DE ICMS',
    'NÃO DESTACAMOS ST', 'TABELA CONSUMO - LINHA HL', 'TABELA REVENDA  - LINHA HL',
    'MATERIAL C/ ANEL', 'TABELA JGS',
}

def _pick_sheet(wb, prefer):
    for name in prefer:
        if name in wb.sheetnames:
            return wb[name]
    return wb[wb.sheetnames[0]]

def _parse_grupos(ws, grupos, tabela_tipo, default_cat):
    """
    Percorre a planilha em 'grupos' de colunas (col_nome, col_preco),
    mantendo a categoria corrente por grupo. Retorna lista de produtos.
    Também coleta categorias que apareceram SEM nenhum item com preço
    (úteis para avisar o usuário).
    """
    produtos = []
    categoria_atual = [None] * len(grupos)
    cat_com_preco = set()
    cat_vistas = set()

    for row in ws.iter_rows(min_row=2, values_only=True):
        for i, (col_nome, col_preco) in enumerate(grupos):
            nome = clean_text(row[col_nome]) if col_nome < len(row) else None
            preco_raw = row[col_preco] if col_preco < len(row) else None

            if not nome or nome in SKIP_TEXTS:
                continue

            try:
                preco = float(preco_raw) if preco_raw is not None else None
            except (ValueError, TypeError):
                preco = None

            if is_header(nome, preco):
                categoria_atual[i] = nome
                cat_vistas.add(nome)
                continue

            if preco is not None and preco > 0:
                cat = categoria_atual[i] or default_cat
                cat_com_preco.add(cat)
                produtos.append({
                    "categoria":  cat,
                    "tamanho":    nome,
                    "descricao":  f"{cat} {nome}".strip(),
                    "preco":      round(preco, 2),
                    "tabela":     tabela_tipo,
                    "ncm":        inferir_ncm(cat),
                })

    sem_preco = sorted(cat_vistas - cat_com_preco)
    return produtos, sem_preco


def parse_hl_table(filepath, tabela_tipo):
    """Tabelas HL (CONSUMO/REVENDA). 3 grupos de colunas [nome, preço, vazio]."""
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = _pick_sheet(wb, ["Revenda"])
    grupos = [(0, 1), (3, 4), (6, 7)]
    return _parse_grupos(ws, grupos, tabela_tipo, default_cat="Geral")


def parse_pressao_table(filepath):
    """
    Tabela de pressão (JGS).
    Apenas a aba 'Plan1' contém preços (4 grupos [vazio,nome,preço]).
    Plan2/Plan3 trazem só listas de tamanhos sem preço — são ignoradas,
    mas seus nomes de família são reportados como aviso.
    """
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = _pick_sheet(wb, ["Plan1"])
    grupos = [(1, 2), (4, 5), (7, 8), (10, 11)]
    produtos, _ = _parse_grupos(ws, grupos, "PRESSAO_JGS", default_cat="JGS")

    # Famílias listadas em Plan2/Plan3 (sem preço) — só para aviso.
    familias_sem_preco = []
    for sn in ("Plan2", "Plan3"):
        if sn in wb.sheetnames:
            wsx = wb[sn]
            for row in wsx.iter_rows(values_only=True):
                for cell in row:
                    nome = clean_text(cell)
                    if nome and is_header(nome, None) and nome not in SKIP_TEXTS \
                       and not re.search(r'\d', nome):  # só famílias "puras"
                        if nome not in familias_sem_preco:
                            familias_sem_preco.append(nome)
    return produtos, familias_sem_preco


def inferir_ncm(categoria):
    """Infere o código NCM com base na categoria do produto."""
    if not categoria:
        return "7307.11.00"
    cat = categoria.upper()
    ncm_map = {
        "TUBINT": "7303.00.00", "TUBO": "7303.00.00",
        "JOELHO": "7307.11.00", "CURVA": "7307.11.00",
        "TEE": "7307.11.00", "TÊ": "7307.11.00",
        "LUVA": "7307.11.00", "JUNÇÃO": "7307.11.00", "JUNCAO": "7307.11.00",
        "JUNTA": "7307.11.00", "BUCHA": "7307.11.00", "PLACA CEGA": "7307.11.00",
        "CRUZETA": "7307.11.00", "REDUÇÃO": "7307.11.00", "RED.": "7307.11.00",
        "RED ": "7307.11.00", "EXTREMIDADE": "7307.11.00", "EXTREM": "7307.11.00",
        "TOCO": "7307.11.00", "CAPS": "7307.11.00", "FLANGE": "7307.11.00",
        "COLAR": "7307.11.00", "ULTRALINK": "7307.11.00", "ULTRAQUIK": "7307.11.00",
        "CJ. DE ACESS": "7307.11.00",
        "ANEL": "4016.93.00", "TAMPÃO": "4016.93.00", "ADAPTADOR BORR": "4016.93.00",
        "VALV": "8481.20.00", "VÁLV": "8481.20.00",
        "TAMPA": "7325.10.00", "GRELHA": "7325.10.00",
        "PLUG": "3926.90.40",
        "HIDRANTE": "8481.80.39", "VENTOSA": "8481.80.39",
        "RALO": "3922.90.00",
        "LUBRIFICANTE": "3403.19.90",
        "PARAFUSO": "7318.15.00",
    }
    for key, ncm in ncm_map.items():
        if key in cat:
            return ncm
    return "7307.11.00"


# ════════════════════════════════════════════════════════════════════════════
# MAIN
# ════════════════════════════════════════════════════════════════════════════

def _achar(*palavras):
    """Acha um .xlsx na pasta do script cujo nome contenha TODAS as palavras."""
    for caminho in glob.glob(os.path.join(SCRIPT_DIR, "*.xlsx")):
        nome = os.path.basename(caminho).upper()
        if all(p.upper() in nome for p in palavras):
            return caminho
    return None

def main():
    todos = []
    avisos = []

    fontes = [
        ("HL CONSUMO", parse_hl_table, ("CONSUMO",), "HL_CONSUMO"),
        ("HL REVENDA", parse_hl_table, ("REVENDA",), "HL_REVENDA"),
        # Pressão JGS agora vem da BASE IA completa (NCM + SAP + preço de venda).
        # Procura "tabela base IA *.xlsx"; se não achar, cai na tabela PRESS antiga.
        ("PRESSÃO JGS (base IA)", "BASE_IA", ("BASE", "IA"), "PRESSAO_JGS"),
        ("SMU", parse_hl_table, ("SMU",), "SMU"),
    ]

    for label, func, palavras, tipo in fontes:
        caminho = _achar(*palavras)
        if not caminho and func == "BASE_IA":
            caminho = _achar("PRESS")  # fallback p/ a tabela de pressão antiga
            func = parse_pressao_table
        if not caminho:
            avisos.append(f"⚠️ Arquivo de {label} não encontrado na pasta.")
            print(f"⚠️  {label}: arquivo não encontrado (procurei por {palavras}).")
            continue
        print(f"Processando {label}: {os.path.basename(caminho)}")
        if func == "BASE_IA":
            from extrair_base_ia import converter
            p = converter(caminho)
        elif func is parse_hl_table:
            p, _ = parse_hl_table(caminho, tipo)
        else:
            p, familias = parse_pressao_table(caminho)
            if familias:
                avisos.append(
                    "ℹ️ Famílias da linha de pressão SEM preço na tabela (não cotáveis): "
                    + ", ".join(familias)
                )
        todos.extend(p)
        print(f"  → {len(p)} produtos")

    out_path = os.path.join(SCRIPT_DIR, "catalogo_produtos.json")
    with open(out_path, "w", encoding="utf-8") as f:
        json.dump(todos, f, ensure_ascii=False, indent=2)

    print(f"\n✅ Total: {len(todos)} produtos")
    print(f"   Salvo em: {out_path}")
    if avisos:
        print("\n--- AVISOS ---")
        for a in avisos:
            print(a)
    return todos


if __name__ == "__main__":
    main()
