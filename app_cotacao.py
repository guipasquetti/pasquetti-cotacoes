"""
app_cotacao.py — Pasquetti Gerador de Cotações
Iniciar: streamlit run app_cotacao.py
"""

import streamlit as st
import json, re, difflib, unicodedata, datetime, io, base64, os
from pathlib import Path
import openpyxl

# Camada de IA (opcional). Se a lib/chave não existirem, o app segue só no
# motor determinístico.
try:
    import matcher_ia
except Exception:
    matcher_ia = None
try:
    import dados_supabase
except Exception:
    dados_supabase = None
try:
    from streamlit_searchbox import st_searchbox
except Exception:
    st_searchbox = None
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# ── Configuração da página ─────────────────────────────────────────────────
st.set_page_config(
    page_title="Pasquetti — Cotações",
    page_icon="🔧",
    layout="wide",
    initial_sidebar_state="expanded",
)

SCRIPT_DIR   = Path(__file__).parent
CATALOG_JSON = SCRIPT_DIR / "catalogo_produtos.json"
LOGO_PATH    = SCRIPT_DIR / "logo_pasquetti.png"

# ── Web: ponte de segredos + tela de senha ──────────────────────────────────
# Em produção (Streamlit Cloud) a chave de IA e a senha vêm de st.secrets.
# Localmente, sem secrets, nada disso atrapalha (o app abre direto).
def _secret(nome):
    try:
        return st.secrets.get(nome)
    except Exception:
        return None

# Disponibiliza a chave de IA do servidor para o matcher_ia (via variável de ambiente)
_chave_servidor = _secret("ANTHROPIC_API_KEY")
if _chave_servidor and not os.environ.get("ANTHROPIC_API_KEY"):
    os.environ["ANTHROPIC_API_KEY"] = _chave_servidor

# Credenciais do Supabase (aprendizado + ST) vindas dos segredos na web
for _k in ("SUPABASE_URL", "SUPABASE_KEY"):
    _v = _secret(_k)
    if _v and not os.environ.get(_k):
        os.environ[_k] = _v

def _checar_senha():
    """Se houver APP_PASSWORD nos segredos, exige senha. Sem senha definida, libera."""
    senha = _secret("APP_PASSWORD")
    if not senha:
        return  # uso local / sem proteção configurada
    if st.session_state.get("_auth_ok"):
        return
    st.markdown("### 🔒 Pasquetti — Sistema de Cotações")
    st.caption("Acesso restrito. Informe a senha para continuar.")
    pw = st.text_input("Senha", type="password", label_visibility="collapsed",
                       placeholder="Senha de acesso")
    if st.button("Entrar", type="primary"):
        if pw == senha:
            st.session_state["_auth_ok"] = True
            st.rerun()
        else:
            st.error("Senha incorreta.")
    st.stop()

_checar_senha()

# ── Feedback persistente (sobrevive ao st.rerun) ────────────────────────────
# Para os vendedores acompanharem cada ação: ao salvar/alterar algo seguido de
# st.rerun(), guardamos a mensagem aqui e ela é exibida no recarregamento.
def flash(msg, kind="success"):
    st.session_state["_flash"] = (kind, msg)

def _render_flash():
    fm = st.session_state.pop("_flash", None)
    if not fm:
        return
    kind, msg = fm
    icon = {"success":"✅","info":"ℹ️","warning":"⚠️","error":"❌"}.get(kind, "ℹ️")
    try:
        st.toast(msg, icon=icon)
    except Exception:
        {"success":st.success,"info":st.info,"warning":st.warning,
         "error":st.error}.get(kind, st.info)(msg)

_render_flash()

# ── Cores da marca ──────────────────────────────────────────────────────────
NAVY   = "#1B3065"   # azul-marinho Pasquetti
COPPER = "#C47A3A"   # cobre/laranja Pasquetti
NAVY_L = "#2a4a8a"   # azul claro hover
COPPER_L = "#e09050" # cobre claro

# ── Logo como SVG inline (ícone "P" estilizado com cano) ──────────────────
LOGO_SVG = """
<svg width="52" height="52" viewBox="0 0 52 52" fill="none" xmlns="http://www.w3.org/2000/svg">
  <!-- Fundo do ícone -->
  <rect width="52" height="52" rx="8" fill="#1B3065"/>
  <!-- Corpo do P (cano vertical) -->
  <rect x="12" y="8" width="9" height="36" rx="3" fill="white"/>
  <!-- Arco superior (curva do cano) -->
  <path d="M21 8 Q38 8 38 20 Q38 32 21 32" stroke="white" stroke-width="9" fill="none" stroke-linecap="round"/>
  <!-- Detalhe cobre - conector -->
  <rect x="9" y="34" width="15" height="8" rx="2" fill="#C47A3A"/>
  <!-- Detalhe cobre - flange superior -->
  <rect x="9" y="6" width="15" height="6" rx="2" fill="#C47A3A"/>
</svg>
"""

# Logos reais (horizontal p/ cabeçalho e Excel; vertical p/ barra lateral).
LOGO_HORIZ = SCRIPT_DIR / "logo_horizontal.png"
LOGO_VERT  = SCRIPT_DIR / "logo_vertical.png"

def _b64_de(path):
    try:
        if path and path.exists():
            return base64.b64encode(path.read_bytes()).decode()
    except Exception:
        pass
    return None

def logo_b64(prefer="horizontal"):
    """Retorna base64 do logo preferido, com fallback para as outras versões."""
    if prefer == "vertical":
        ordem = [LOGO_VERT, LOGO_HORIZ, LOGO_PATH]
    else:
        ordem = [LOGO_HORIZ, LOGO_PATH, LOGO_VERT]
    for p in ordem:
        b = _b64_de(p)
        if b:
            return b
    return None

def caminho_logo_excel():
    """Caminho do melhor logo para o Excel (horizontal de preferência)."""
    for p in (LOGO_HORIZ, LOGO_PATH, LOGO_VERT):
        if p.exists():
            return str(p)
    return None

# Compatibilidade
def get_logo_b64():
    return logo_b64("horizontal")

# ════════════════════════════════════════════════════════════════════════════
# CSS — Identidade Visual Pasquetti
# ════════════════════════════════════════════════════════════════════════════
st.markdown(f"""
<style>
  @import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700&display=swap');

  html, body, .stApp {{
    font-family: 'Inter', sans-serif !important;
    background: #f0f3f9 !important;
  }}

  /* ── Sidebar ── */
  [data-testid="stSidebar"] {{
    background: linear-gradient(180deg, {NAVY} 0%, #16233f 100%) !important;
    border-right: 3px solid {COPPER} !important;
  }}
  [data-testid="stSidebar"] * {{ color: #e3ecff !important; }}
  [data-testid="stSidebar"] strong {{ color: #ffffff !important; }}
  /* Cabeçalhos de seção: barra cobre + caixa-alta */
  [data-testid="stSidebar"] h2,
  [data-testid="stSidebar"] h3 {{
    color: #ffffff !important;
    font-size: 12.5px !important; font-weight: 700 !important;
    text-transform: uppercase; letter-spacing: .7px;
    border-left: 3px solid {COPPER}; padding: 2px 0 2px 10px !important;
    margin: 4px 0 12px 0 !important;
  }}
  /* Campos de texto / número */
  [data-testid="stSidebar"] .stTextInput input,
  [data-testid="stSidebar"] .stNumberInput input {{
    background: rgba(255,255,255,0.10) !important;
    color: #ffffff !important;
    border: 1px solid rgba(255,255,255,0.22) !important;
    border-radius: 8px !important;
    padding: 9px 12px !important;
  }}
  [data-testid="stSidebar"] .stTextInput input:focus,
  [data-testid="stSidebar"] .stNumberInput input:focus {{
    border-color: {COPPER} !important;
    box-shadow: 0 0 0 2px rgba(199,122,52,0.25) !important;
  }}
  /* Selectbox (baseweb) sobre fundo escuro */
  [data-testid="stSidebar"] [data-baseweb="select"] > div {{
    background: rgba(255,255,255,0.10) !important;
    border: 1px solid rgba(255,255,255,0.22) !important;
    border-radius: 8px !important;
    color: #ffffff !important;
  }}
  [data-testid="stSidebar"] [data-baseweb="select"] svg {{ fill: #e3ecff !important; }}
  [data-testid="stSidebar"] label {{ color: #aec4f2 !important; font-size:12px !important; font-weight:600 !important; }}
  [data-testid="stSidebar"] hr {{ border-color: rgba(255,255,255,0.14) !important; margin: 14px 0 !important; }}
  /* Botões da barra lateral: legíveis sobre o fundo escuro */
  [data-testid="stSidebar"] div.stButton > button {{
    background: rgba(255,255,255,0.10) !important;
    color: #ffffff !important;
    border: 1px solid {COPPER} !important;
    border-radius: 8px !important;
    font-weight: 600 !important;
  }}
  [data-testid="stSidebar"] div.stButton > button:hover {{
    background: {COPPER} !important;
    border-color: {COPPER} !important;
  }}

  /* ── Header ── */
  .pq-header {{
    background: linear-gradient(135deg, {NAVY} 0%, #243d7a 100%);
    border-radius: 12px;
    padding: 20px 28px;
    margin-bottom: 20px;
    display: flex;
    align-items: center;
    gap: 16px;
    border-bottom: 4px solid {COPPER};
  }}
  .pq-title {{ color: #fff; font-size: 22px; font-weight: 700; margin: 0; letter-spacing: -0.3px; }}
  .pq-sub   {{ color: #b8cfff; font-size: 12px; margin: 3px 0 0 0; }}
  .pq-tag   {{
    background: {COPPER}; color: white;
    font-size: 10px; font-weight: 600;
    padding: 3px 8px; border-radius: 20px;
    text-transform: uppercase; letter-spacing: .5px;
    margin-left: auto;
  }}

  /* ── Tabs ── */
  .stTabs [data-baseweb="tab-list"] {{
    background: white !important;
    border-radius: 8px !important;
    padding: 4px !important;
    border: 1px solid #dce3f0 !important;
    gap: 2px !important;
  }}
  .stTabs [data-baseweb="tab"] {{
    border-radius: 6px !important;
    padding: 8px 16px !important;
    font-size: 13px !important;
    font-weight: 500 !important;
    color: #5a6a8a !important;
  }}
  .stTabs [aria-selected="true"] {{
    background: {NAVY} !important;
    color: white !important;
  }}

  /* ── Botão principal ── */
  div.stButton > button[kind="primary"] {{
    background: linear-gradient(135deg, {COPPER}, #e09050) !important;
    color: white !important;
    border: none !important;
    padding: 13px 28px !important;
    font-size: 15px !important;
    font-weight: 600 !important;
    border-radius: 8px !important;
    width: 100% !important;
    letter-spacing: 0.3px !important;
  }}
  div.stButton > button[kind="primary"]:hover {{ opacity: 0.9 !important; }}

  /* ── Botão secundário (CNPJ) ── */
  div.stButton > button:not([kind="primary"]) {{
    background: {NAVY} !important;
    color: white !important;
    border: none !important;
    border-radius: 6px !important;
    font-size: 12px !important;
    font-weight: 600 !important;
    padding: 6px 14px !important;
  }}

  /* ── Cards de resultado ── */
  .pq-metric {{
    background: white;
    border-radius: 10px;
    padding: 16px;
    text-align: center;
    border: 1px solid #dce3f0;
    border-top: 4px solid {NAVY};
  }}
  .pq-metric.copper {{ border-top-color: {COPPER}; }}
  .pq-metric.red    {{ border-top-color: #c0392b; }}
  .pq-metric.green  {{ border-top-color: #27ae60; }}
  .pq-metric-val  {{ font-size: 28px; font-weight: 700; color: {NAVY}; }}
  .pq-metric-val.copper {{ color: {COPPER}; }}
  .pq-metric-val.red    {{ color: #c0392b; }}
  .pq-metric-val.green  {{ color: #27ae60; }}
  .pq-metric-lbl  {{ font-size: 11px; color: #7a8aaa; margin-top: 4px; font-weight: 500; }}

  /* ── Tabela de itens ── */
  .pq-table {{ width:100%; border-collapse:collapse; font-size:12px; }}
  .pq-table th {{
    background: {NAVY}; color: white;
    padding: 9px 12px; text-align: left;
    font-weight: 600; font-size: 11px; text-transform: uppercase; letter-spacing:.4px;
  }}
  .pq-table td {{ padding: 9px 12px; border-bottom: 1px solid #eef1f8; vertical-align:middle; }}
  .pq-table tr:nth-child(even) td {{ background: #f7f9ff; }}
  .pq-table tr:hover td {{ background: #eef3ff; }}
  .badge {{ border-radius:4px; padding:2px 8px; font-size:10px; font-weight:700; }}
  .badge-a  {{ background:#e8f5e9; color:#2e7d32; }}
  .badge-m  {{ background:#fff8e1; color:#c77a00; }}
  .badge-s  {{ background:#fce4ec; color:#b71c1c; }}

  /* ── Seção de sugestões ── */
  .pq-suggest {{
    background: #fffbf3;
    border: 1px solid #f0d090;
    border-left: 4px solid {COPPER};
    border-radius: 8px;
    padding: 14px 18px;
    margin-top: 12px;
    font-size: 12px;
  }}
  .pq-suggest-title {{ font-weight: 700; color: {COPPER}; margin-bottom: 8px; font-size:13px; }}

  /* ── Seção não encontrados ── */
  .pq-notfound {{
    background: #fff5f5;
    border: 1px solid #f0c0c0;
    border-left: 4px solid #c0392b;
    border-radius: 8px;
    padding: 14px 18px;
    margin-top: 12px;
    font-size: 12px;
  }}
  .pq-notfound-title {{ font-weight: 700; color: #c0392b; margin-bottom: 8px; font-size:13px; }}

  /* ── Upload area ── */
  [data-testid="stFileUploader"] {{
    background: white !important;
    border: 2px dashed #b0bdd8 !important;
    border-radius: 10px !important;
    padding: 20px !important;
  }}

  /* ── Seção títulos ── */
  .pq-section {{
    font-size: 14px; font-weight: 700;
    color: {NAVY};
    border-bottom: 2px solid {COPPER};
    padding-bottom: 6px;
    margin: 16px 0 12px 0;
    display: inline-block;
  }}

  /* ── CNPJ strip ── */
  .cnpj-strip {{
    background: white;
    border: 1px solid #dce3f0;
    border-radius: 8px;
    padding: 10px 14px;
    margin-bottom: 10px;
    font-size: 11px;
  }}
  .cnpj-label {{ color: #7a8aaa; font-size: 10px; text-transform:uppercase; letter-spacing:.4px; }}

  #MainMenu, footer, [data-testid="stToolbar"] {{ visibility: hidden !important; }}

  /* ── Campos de entrada: contraste no fundo claro (regra global) ── */
  .stApp input, .stApp textarea,
  .stApp [data-baseweb="input"] input,
  .stApp [data-baseweb="base-input"] input,
  .stApp [data-baseweb="textarea"] textarea {{
    color: {NAVY} !important;
    background: #ffffff !important;
    -webkit-text-fill-color: {NAVY} !important;
  }}
  .stApp input::placeholder,
  .stApp textarea::placeholder,
  .stApp [data-baseweb="input"] input::placeholder,
  .stApp [data-baseweb="base-input"] input::placeholder,
  .stApp [data-baseweb="textarea"] textarea::placeholder {{
    color: #556682 !important;
    -webkit-text-fill-color: #556682 !important;   /* cinza-azulado ESCURO, alto contraste */
    opacity: 1 !important;
  }}

  /* ── Reafirma os campos da BARRA LATERAL (fundo escuro) ── */
  [data-testid="stSidebar"] input,
  [data-testid="stSidebar"] textarea {{
    color: #ffffff !important;
    -webkit-text-fill-color: #ffffff !important;
    background: rgba(255,255,255,0.12) !important;
  }}
  [data-testid="stSidebar"] input::placeholder,
  [data-testid="stSidebar"] textarea::placeholder {{
    color: #c4d4f5 !important;
    -webkit-text-fill-color: #c4d4f5 !important;
    opacity: 1 !important;
  }}

  /* ── Texto dos expanders ("Ver itens lidos") legível no fundo claro ── */
  [data-testid="stExpander"] summary,
  [data-testid="stExpander"] p,
  [data-testid="stExpander"] li,
  [data-testid="stExpander"] span,
  [data-testid="stExpander"] div {{
    color: {NAVY} !important;
  }}
  /* mantém os expanders da barra lateral (fundo escuro) com texto claro */
  [data-testid="stSidebar"] [data-testid="stExpander"] summary,
  [data-testid="stSidebar"] [data-testid="stExpander"] p,
  [data-testid="stSidebar"] [data-testid="stExpander"] li,
  [data-testid="stSidebar"] [data-testid="stExpander"] span,
  [data-testid="stSidebar"] [data-testid="stExpander"] div {{
    color: #d8e4ff !important;
  }}
</style>
""", unsafe_allow_html=True)


# ════════════════════════════════════════════════════════════════════════════
# MOTOR DE CORRESPONDÊNCIA
# ════════════════════════════════════════════════════════════════════════════

SINONIMOS = {
    r'\bjf\b':'joelho fofo', r'\bjff\b':'joelho fofo',
    r'\bjoelho\s*fofo\b':'joelho fofo', r'\bjoelho\b':'joelho fofo',
    r'\bcurva\s*fofo\b':'joelho fofo', r'\bcurva\b':'joelho fofo',
    r'\btee\s*fofo\b':'tee fofo', r'\btê\b':'tee fofo',
    r'\btee\b':'tee fofo', r'\bte\b':'tee fofo',
    r'\bluva\s*fofo\b':'luva fofo', r'\bluva\s*hl\b':'luva fofo',
    r'\bljf\b':'luva fofo',
    r'\bbucha\s*red\b':'bucha reducao', r'\bbucha\b':'bucha reducao',
    r'\bred\s*conc\b':'reducao', r'\bred\b':'reducao',
    r'\bpl\s*cega\b':'placa cega fofo', r'\bplaca\s*cega\b':'placa cega fofo',
    r'\bjuncao\b':'juncao fofo', r'\bjunc\b':'juncao fofo',
    r'\banel\s*borr\b':'anel borr', r'\banel\s*borracha\b':'anel borr',
    r'\banel\b':'anel borr',
    r'\bjunta\s*rap\b':'junta rapid', r'\bjr\b':'junta rapid jrsmu',
    r'\bjrsmu\b':'junta rapid jrsmu',
    r'\btampao\b':'tampao borr',
    r'\bgrelha\b':'grelha hemisf fofo',
    r'\bvalvula\b':'valv', r'\bvalv\s*gav\b':'valv gav',
    r'\bvalv\s*ret\b':'valv ret', r'\bvalv\b':'valv',
    r'\bhidrante\b':'hidrante col',
    r'\bventosa\b':'ventosa', r'\bvsr\b':'ventosa vsr',
    r'\bvsf\b':'ventosa vsf', r'\bvtf\b':'ventosa vtf',
    r'\bcolar\b':'colar de tomada',
    r'\banel\s*trav\b':'anel trav interno',
    r'\bjtd\b':'junta desmontagem', r'\bjg\b':'junta gibault',
    r'\bul\b':'ultralink', r'\buq\b':'ultraquik',
    r'\bflange\s*av\b':'flange avulsa', r'\bflange\s*cego\b':'flange cego',
    r'\btoco\b':'toco c/fl', r'\bluva\s*cr\b':'luva cr jgs',
    r'\bluva\s*jm\b':'luva junta mecanica', r'\bluva\s*gt\b':'luva gr tol',
    r'\bluva\s*trip\b':'luva tripartida', r'\bcruz\b':'cruzeta',
    r'\bbb\b':'b/b', r'\bpp\b':'p/p', r'\bpb\b':'p/b',
    r'\bbfl\b':'b/fl', r'\bcfl\b':'c/fl', r'\bpfl\b':'p/fl',
    r'\b90\s*gr(?:aus?)?\b':'90', r'\b45\s*gr(?:aus?)?\b':'45',
    r'\b90°':'90', r'\b45°':'45',
    r'\b(\d+)\s*mm\b':r'\1mm', r'\b(\d+)\s*x\s*(\d+)\b':r'\1x\2',
    r'\bhl\b':'hl', r'\bjgs\b':'jgs', r'\bsmu\b':'smu', r'\bsme\b':'sme',
}

# Palavras-chave de tipo de produto para bônus de categoria
TIPO_KEYWORDS = {
    'joelho fofo': ['joelho','jf','jff','curva'],
    'tee fofo':    ['tee','te'],
    'luva fofo':   ['luva'],
    'anel borr':   ['anel','borracha'],
    'tubo':        ['tubo','tub'],
    'junta rapid': ['junta','jr','jrsmu'],
    'bucha':       ['bucha','reducao','red'],
    'placa cega':  ['placa','cega'],
    'valv gav':    ['valvula','valv','gav'],
    'valv ret':    ['valv','ret'],
    'hidrante':    ['hidrante'],
    'ventosa':     ['ventosa','vsr','vsf','vtf'],
    'flange':      ['flange'],
}

ANGULOS = {"90", "45", "22", "11", "15"}

def remover_acentos(t):
    return ''.join(c for c in unicodedata.normalize('NFD', t)
                   if unicodedata.category(c) != 'Mn')

def normalizar(texto):
    if not texto: return ""
    t = remover_acentos(str(texto)).lower()
    t = re.sub(r'[^\w\s./x-]', ' ', t)
    return re.sub(r'\s+', ' ', t).strip()

def expandir(texto):
    t = normalizar(texto)
    for p, s in SINONIMOS.items():
        t = re.sub(p, s, t, flags=re.IGNORECASE)
    return re.sub(r'\s+', ' ', t).strip()

def _fmt_cnpj(cnpj):
    d = "".join(ch for ch in str(cnpj or "") if ch.isdigit())
    if len(d) != 14: return cnpj or ""
    return f"{d[:2]}.{d[2:5]}.{d[5:8]}/{d[8:12]}-{d[12:]}"

def _fill_cliente(r):
    """Preenche os campos do cliente na sessão a partir de um registro (base ou SEFAZ)."""
    st.session_state["input_nome"]     = r.get("nome","") or ""
    st.session_state["input_fantasia"] = r.get("fantasia","") or ""
    st.session_state["input_end"]      = r.get("endereco","") or ""
    st.session_state["input_tel"]      = r.get("telefone","") or ""
    st.session_state["input_email"]    = r.get("email","") or ""
    st.session_state["input_cnpj"]     = _fmt_cnpj(r.get("cnpj",""))
    if r.get("uf"): st.session_state["uf_sel"] = r["uf"]

def cond_similar(texto, lista, limiar=0.7):
    """Condição cadastrada mais parecida (ratio>=limiar) ou None."""
    alvo = normalizar(texto); melhor, mr = None, 0.0
    for c in lista:
        r = difflib.SequenceMatcher(None, alvo, normalizar(c)).ratio()
        if r > mr: mr, melhor = r, c
    return melhor if (melhor and mr >= limiar) else None

def extrair_diametros(texto):
    t = normalizar(texto)
    result = set()
    for m in re.finditer(r'(\d+)x(\d+)', t): result.add(m.group(0))
    for m in re.finditer(r'(\d+)\s*mm', t):
        if m.group(1) not in ANGULOS: result.add(m.group(1))
    for m in re.finditer(r'\b(\d+)\b', t):
        n = m.group(1)
        if n not in ANGULOS and 40 <= int(n) <= 600: result.add(n)
    return result

def bonus_tipo(query_exp, cat_desc):
    """Bônus se query e produto são do mesmo tipo."""
    for tipo, keywords in TIPO_KEYWORDS.items():
        tipo_norm = normalizar(tipo)
        if any(k in query_exp for k in keywords):
            if tipo_norm in cat_desc:
                return 12
    return 0

def score_ts(a, b):
    ta = " ".join(sorted(a.split()))
    tb = " ".join(sorted(b.split()))
    return difflib.SequenceMatcher(None, ta, tb).ratio() * 100

def score_set(a, b):
    sa, sb = set(a.split()), set(b.split())
    inter = sa & sb
    t0 = " ".join(sorted(inter))
    t1 = " ".join(sorted(inter | (sa - inter)))
    t2 = " ".join(sorted(inter | (sb - inter)))
    return max(difflib.SequenceMatcher(None, t0, t1).ratio(),
               difflib.SequenceMatcher(None, t0, t2).ratio(),
               difflib.SequenceMatcher(None, t1, t2).ratio()) * 100

def buscar(descricao, catalogo):
    """
    Retorna (produto, score, confiança).
    CONFIRMADO: score ≥ 80%  → entra na cotação automaticamente
    SUGESTÃO:   score 60-79% → listado para revisão manual, NÃO entra na cotação
    NÃO ENCONTRADO: < 60%
    """
    qo = normalizar(descricao)
    qe = expandir(descricao)
    best_prod, best_score = None, 0.0

    for p in catalogo:
        d = p["_norm"]
        # Múltiplas estratégias
        s = max(
            score_ts(qe, d),
            score_set(qe, d),
            score_ts(qo, d) * 0.9,
            score_set(qo, d) * 0.9,
        )
        # Bônus por tipo de produto
        s += bonus_tipo(qe, d)
        # Bônus/penalidade por diâmetro
        dq, dp = extrair_diametros(qe), extrair_diametros(d)
        if dq and dp:
            s += 18 if dq & dp else -28
        s = min(100, max(0, s))

        if s > best_score:
            best_score, best_prod = s, p

    if best_score >= 80:
        return best_prod, best_score, "CONFIRMADO"
    elif best_score >= 60:
        return best_prod, best_score, "SUGESTÃO"
    return best_prod, best_score, "NÃO ENCONTRADO"


def processar(itens_brutos, catalogo):
    confirmados, sugestoes, nao_encontrados = [], [], []
    for item in itens_brutos:
        p, score, conf = buscar(item["descricao"], catalogo)
        qtd = item["quantidade"]
        base = {**item, "score": score, "conf": conf}
        if p:
            base.update({"produto": p["descricao"], "ncm": p.get("ncm",""),
                          "preco": p["preco"], "total": qtd * p["preco"]})
        if conf == "CONFIRMADO":
            confirmados.append(base)
        elif conf == "SUGESTÃO":
            sugestoes.append(base)
        else:
            nao_encontrados.append(base)
    return confirmados, sugestoes, nao_encontrados


def _hint_deterministico(descricao, catalogo, n=3):
    """Top-N candidatos do motor determinístico, como dica para a IA."""
    qe = expandir(descricao)
    scored = []
    for p in catalogo:
        s = max(score_ts(qe, p["_norm"]), score_set(qe, p["_norm"]))
        scored.append((s, p["descricao"]))
    scored.sort(reverse=True)
    return " | ".join(d for _, d in scored[:n])


def processar_hibrido(itens_brutos, catalogo, usar_ia, correcoes=None, nao_trabalhados=None):
    """
    Aprendizado (correções salvas) + motor determinístico + (opcional) IA.
    Prioridade: 0) item marcado como 'não trabalhamos' → 1) correção aprendida
    (exata) → 2) IA → 3) determinístico.
    Retorna (confirmados, sugestoes, nao_encontrados, nao_trabalhados).
    """
    correcoes = correcoes or {}
    nao_trabalhados = nao_trabalhados or set()
    por_desc = {p["descricao"]: p for p in catalogo}  # produto por descrição

    # 1) Determinístico para todos (sempre)
    det = []
    for item in itens_brutos:
        p, score, conf = buscar(item["descricao"], catalogo)
        det.append({"item": item, "prod": p, "score": score, "conf": conf})

    ia_res = None
    if usar_ia and matcher_ia is not None:
        try:
            hints = [_hint_deterministico(it["descricao"], catalogo) for it in itens_brutos]
            ia_res = matcher_ia.interpretar_itens(itens_brutos, catalogo, hints=hints, correcoes=correcoes)
        except Exception as e:
            st.warning(f"IA indisponível nesta cotação ({e}). Usando apenas o motor determinístico.")
            ia_res = None

    confirmados, sugestoes, nao, nao_trab = [], [], [], []
    for i, item in enumerate(itens_brutos):
        qtd = item["quantidade"]

        # 0) Item que NÃO trabalhamos: sai da cotação, não vai para o Excel.
        if normalizar(item["descricao"]) in nao_trabalhados:
            nao_trab.append({"descricao": item["descricao"], "quantidade": qtd,
                             "conf": "NÃO TRABALHAMOS"})
            continue

        d = det[i]

        # Resultado determinístico como padrão
        prod, score, conf, fonte, just = d["prod"], d["score"], d["conf"], "auto", ""

        # 1) Correção aprendida tem prioridade máxima (match exato normalizado)
        chave = normalizar(item["descricao"])
        if chave in correcoes and correcoes[chave] in por_desc:
            prod = por_desc[correcoes[chave]]
            score, conf, fonte, just = 100.0, "CONFIRMADO", "aprendizado", "Correção aprendida"
            base = {"descricao": item["descricao"], "quantidade": qtd,
                    "score": score, "conf": conf, "fonte": fonte, "justificativa": just,
                    "produto": prod["descricao"], "ncm": prod.get("ncm", ""),
                    "preco": prod["preco"], "total": qtd * prod["preco"]}
            confirmados.append(base)
            continue

        if ia_res is not None:
            r = ia_res[i]
            ia_prod = catalogo[r["indice"]] if r["indice"] is not None else None
            # IA é primária quando achou algo
            if ia_prod is not None:
                prod, score, conf = ia_prod, r["confianca"], r["status"]
                fonte, just = "IA", r.get("justificativa", "")
            # Rede de segurança: IA não achou, mas determinístico tem match exato forte
            elif d["conf"] == "CONFIRMADO":
                prod, score, conf, fonte = d["prod"], d["score"], "CONFIRMADO", "auto"
            else:
                # Mantém o melhor palpite do motor determinístico como sugestão,
                # para que NENHUM item fique sem candidato na lista de aprendizado.
                prod = d["prod"]
                score, conf = d["score"], "NÃO ENCONTRADO"
                just = r.get("justificativa", "") or "Melhor palpite (revisar)"

        base = {"descricao": item["descricao"], "quantidade": qtd,
                "score": score, "conf": conf, "fonte": fonte, "justificativa": just}
        if prod:
            base.update({"produto": prod["descricao"], "ncm": prod.get("ncm", ""),
                         "preco": prod["preco"], "total": qtd * prod["preco"]})
        if conf == "CONFIRMADO":
            confirmados.append(base)
        elif conf == "SUGESTÃO":
            sugestoes.append(base)
        else:
            nao.append(base)
    return confirmados, sugestoes, nao, nao_trab


# ════════════════════════════════════════════════════════════════════════════
# BUSCA CNPJ (BrasilAPI — SEFAZ)
# ════════════════════════════════════════════════════════════════════════════

def _cnpj_valido(c):
    """Valida os dígitos verificadores de um CNPJ (14 dígitos)."""
    if len(c) != 14 or c == c[0] * 14:
        return False
    def _dv(nums, pesos):
        r = sum(int(n) * p for n, p in zip(nums, pesos)) % 11
        return "0" if r < 2 else str(11 - r)
    d1 = _dv(c[:12], [5,4,3,2,9,8,7,6,5,4,3,2])
    d2 = _dv(c[:12] + d1, [6,5,4,3,2,9,8,7,6,5,4,3,2])
    return c[12] == d1 and c[13] == d2

def _http_json(url, timeout=8):
    import urllib.request, ssl
    ctx = ssl.create_default_context()
    req = urllib.request.Request(url, headers={
        "User-Agent": "PasquettiCotacoes/1.0", "Accept": "application/json"})
    with urllib.request.urlopen(req, timeout=timeout, context=ctx) as r:
        return json.loads(r.read().decode())

def _fmt_cnpj_mask(cnpj):
    return f"{cnpj[:2]}.{cnpj[2:5]}.{cnpj[5:8]}/{cnpj[8:12]}-{cnpj[12:]}"

def _parse_brasilapi(data, cnpj):
    tel = (data.get("ddd_telefone_1","") or "").strip()
    if tel and len(tel) >= 8:
        tel = f"({tel[:2]}) {tel[2:]}" if len(tel) >= 10 else tel
    end_parts = [data.get("logradouro",""), data.get("numero",""),
                 data.get("complemento",""), data.get("bairro",""),
                 data.get("municipio",""), data.get("uf","")]
    endereco = " ".join(p for p in end_parts if p and str(p).strip())
    cep = data.get("cep","")
    if cep: endereco += f" — CEP: {cep}"
    return {
        "nome":      data.get("razao_social",""),
        "fantasia":  data.get("nome_fantasia",""),
        "cnpj_fmt":  _fmt_cnpj_mask(cnpj),
        "endereco":  endereco,
        "uf":        (data.get("uf","") or "").strip().upper(),
        "telefone":  tel,
        "email":     data.get("email",""),
        "situacao":  data.get("descricao_situacao_cadastral",""),
        "atividade": data.get("cnae_fiscal_descricao",""),
    }

def _parse_receitaws(data, cnpj):
    if str(data.get("status","")).upper() == "ERROR":
        return None
    end_parts = [data.get("logradouro",""), data.get("numero",""),
                 data.get("complemento",""), data.get("bairro",""),
                 data.get("municipio",""), data.get("uf","")]
    endereco = " ".join(p for p in end_parts if p and str(p).strip())
    cep = (data.get("cep","") or "").strip()
    if cep: endereco += f" — CEP: {cep}"
    atv = data.get("atividade_principal") or []
    atividade = atv[0].get("text","") if atv and isinstance(atv[0], dict) else ""
    return {
        "nome":      data.get("nome",""),
        "fantasia":  data.get("fantasia",""),
        "cnpj_fmt":  _fmt_cnpj_mask(cnpj),
        "endereco":  endereco,
        "uf":        (data.get("uf","") or "").strip().upper(),
        "telefone":  (data.get("telefone","") or "").split("/")[0].strip(),
        "email":     data.get("email",""),
        "situacao":  data.get("situacao",""),
        "atividade": atividade,
    }

def buscar_cnpj(cnpj_raw: str):
    """Consulta dados da empresa pelo CNPJ. Tenta BrasilAPI e, se falhar
    (ex.: 404 intermitente), cai para a ReceitaWS antes de desistir."""
    cnpj = re.sub(r'\D', '', cnpj_raw or "")
    if len(cnpj) != 14:
        return None, "CNPJ deve ter 14 dígitos."
    if not _cnpj_valido(cnpj):
        return None, "CNPJ inválido — confira os dígitos digitados."
    erros = []
    # 1) BrasilAPI (fonte primária)
    try:
        data = _http_json(f"https://brasilapi.com.br/api/cnpj/v1/{cnpj}", timeout=8)
        return _parse_brasilapi(data, cnpj), None
    except Exception as e:
        erros.append(f"BrasilAPI: {e}")
    # 2) ReceitaWS (fallback)
    try:
        data = _http_json(f"https://receitaws.com.br/v1/cnpj/{cnpj}", timeout=12)
        parsed = _parse_receitaws(data, cnpj)
        if parsed:
            return parsed, None
        erros.append("ReceitaWS: CNPJ não encontrado")
    except Exception as e:
        erros.append(f"ReceitaWS: {e}")
    blob = " | ".join(erros)
    low = blob.lower()
    if "404" in blob or "não encontrado" in low or "not found" in low:
        return None, "CNPJ não encontrado nas bases públicas. Confira o número ou preencha manualmente."
    if "timeout" in low or "urlopen" in low or "getaddrinfo" in low:
        return None, "Sem conexão para consultar o CNPJ agora. Preencha os dados manualmente."
    return None, f"Não foi possível consultar agora — preencha manualmente. ({blob[:140]})"


# ════════════════════════════════════════════════════════════════════════════
# PARSING DE ENTRADA
# ════════════════════════════════════════════════════════════════════════════

def extrair_qtd_desc(texto):
    texto = texto.strip()
    m = re.match(r'^(\d+(?:[.,]\d+)?)\s*(?:x|un|pç|pc|pcs|und|unid)?\s+(.+)', texto, re.I)
    if m:
        try: qtd = float(m.group(1).replace(",","."))
        except: qtd = 1
        return qtd, m.group(2).strip()
    m = re.search(r'\s+(\d+(?:[.,]\d+)?)\s*(?:x|un|pç|pc|pcs|und|unid)?$', texto, re.I)
    if m:
        try: qtd = float(m.group(1).replace(",","."))
        except: qtd = 1
        return qtd, texto[:m.start()].strip()
    return 1, texto

def parse_texto(texto):
    itens = []
    for linha in texto.splitlines():
        linha = linha.strip()
        if not linha or linha.startswith("#"): continue
        qtd, desc = extrair_qtd_desc(linha)
        if desc and desc.lower() not in ("none","nan"): itens.append({"descricao": desc, "quantidade": qtd})
    return itens

def _abrir_workbook(data):
    """Abre .xlsx com openpyxl; se for .xls antigo, tenta via pandas/xlrd.
    Levanta ValueError com mensagem amigável se não der."""
    try:
        return openpyxl.load_workbook(io.BytesIO(data), data_only=True).active, None
    except Exception as e_openpyxl:
        # Fallback p/ .xls (formato binário antigo) usando pandas
        try:
            import pandas as pd
            df = pd.read_excel(io.BytesIO(data), header=None)
            return ("__df__", df)
        except Exception:
            raise ValueError(
                "Não consegui abrir a planilha. Se for um arquivo .xls antigo, "
                "abra no Excel e salve como .xlsx (Excel Workbook) e tente de novo. "
                f"(detalhe técnico: {e_openpyxl})")

def _detectar_colunas(linhas):
    """Recebe lista de listas (primeiras linhas) e devolve (col_desc, col_qtd, header_row).
    Prioriza a linha que tem descrição E quantidade (cabeçalho real); só então
    aceita uma linha com apenas descrição."""
    so_desc = None
    for i, row in enumerate(linhas[:15], 1):
        row_s = [str(c).lower().strip() if c is not None else "" for c in row]
        cd = cq = None
        for j, cell in enumerate(row_s):
            if any(k in cell for k in ["produto","descri","item","material","especif","peça","peca"]):
                cd = j
            if any(k in cell for k in ["qtd","quant","qt ","pcs","unid","qnt"]):
                cq = j
        if cd is not None and cq is not None:
            return cd, cq, i  # cabeçalho completo: melhor escolha
        if cd is not None and so_desc is None:
            so_desc = (cd, cd + 1, i)  # guarda como plano B
    if so_desc is not None:
        return so_desc
    return 0, 1, 0  # sem cabeçalho reconhecido: col A=descrição, B=qtd, sem linha p/ pular

def parse_xlsx_bytes(data):
    ws, df = _abrir_workbook(data)
    # Caminho pandas (.xls)
    if ws == "__df__":
        linhas = df.values.tolist()
        col_desc, col_qtd, header_row = _detectar_colunas(linhas)
        corpo = linhas[header_row:]
    else:
        todas = list(ws.iter_rows(values_only=True))
        col_desc, col_qtd, header_row = _detectar_colunas(todas)
        corpo = todas[header_row:]
    itens = []
    for row in corpo:
        row = list(row)
        desc = str(row[col_desc]).strip() if col_desc < len(row) and row[col_desc] is not None else ""
        try: qtd = float(row[col_qtd]) if col_qtd < len(row) and row[col_qtd] not in (None,"") else 1
        except (ValueError, TypeError): qtd = 1
        if desc and desc.lower() not in ("none","nan",""):
            itens.append({"descricao": desc, "quantidade": qtd})
    return itens

def parse_imagem_bytes(data):
    try:
        import pytesseract
        from PIL import Image
        img = Image.open(io.BytesIO(data))
        texto = pytesseract.image_to_string(img, lang="por")
        return parse_texto(texto)
    except ImportError:
        st.error("OCR não disponível. Instale: `pip install pytesseract pillow` e o Tesseract OCR.")
        return []


# ════════════════════════════════════════════════════════════════════════════
# CATÁLOGO
# ════════════════════════════════════════════════════════════════════════════

@st.cache_data
def carregar_catalogo(tabela):
    if not CATALOG_JSON.exists(): return []
    with open(CATALOG_JSON, encoding="utf-8") as f: todos = json.load(f)
    mapa = {"consumo":["HL_CONSUMO"],"revenda":["HL_REVENDA"],
            "pressao":["PRESSAO_JGS"],"smu":["SMU"],
            "todos":["HL_CONSUMO","HL_REVENDA","PRESSAO_JGS","SMU"]}
    catalogo = [p for p in todos if p["tabela"] in mapa.get(tabela,["HL_CONSUMO"])]
    for p in catalogo: p["_norm"] = expandir(p["descricao"])
    return catalogo


# ════════════════════════════════════════════════════════════════════════════
# GERADOR DE XLSX
# ════════════════════════════════════════════════════════════════════════════

def bd(e="thin"):
    s = Side(style=e)
    return Border(left=s,right=s,top=s,bottom=s)
def bb():
    return Border(bottom=Side(style="thin"))

def gerar_xlsx_bytes(confirmados, sugestoes_manuais, nao_enc, num, cliente, tabela_nome, vendedor="", condicao_pagamento="À vista", consumidor_final=False):
    _imp_lbl = "DIFAL" if consumidor_final else "ICMS ST"
    wb = openpyxl.Workbook()
    ws = wb.active; ws.title = "Orçamento"
    COR_H, COR_L = "1B3065", "E8EEF8"
    COR_COPPER = "C47A3A"

    for i, w in enumerate([5,44,14,10,13,10,8,13,15],1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Logo no canto superior esquerdo (sobre as colunas A–B)
    _logo_path = caminho_logo_excel()
    if _logo_path:
        try:
            from openpyxl.drawing.image import Image as XLImage
            img = XLImage(_logo_path)
            ratio = (img.height / img.width) if img.width else 0.3
            img.width = 260
            img.height = int(260 * ratio)
            ws.add_image(img, "A2")
        except Exception:
            pass

    ROW = 2
    ws.row_dimensions[ROW].height = 32
    ws.merge_cells(f"C{ROW}:I{ROW}")
    c = ws[f"C{ROW}"]; c.value = "PASQUETTI COMERCIO DE MATERIAIS HIDRÁULICOS LTDA"
    c.font = Font(name="Calibri",bold=True,size=15,color=COR_H)
    c.alignment = Alignment(horizontal="right",vertical="center")

    ROW+=1; ws.row_dimensions[ROW].height = 13
    ws.merge_cells(f"C{ROW}:I{ROW}"); c = ws[f"C{ROW}"]
    c.value = "www.pasquetti.com.br  |  tubos e conexões de ferro fundido"
    c.font = Font(name="Calibri",size=9,color="555555")
    c.alignment = Alignment(horizontal="right",vertical="center")

    for txt in ["CNPJ: 48.509.178/0001-99  |  Inscrição Estadual: 109.769.567.118",
                "RUA ARAPIRANGA, 55 - 93  |  VILA FORMOSA  |  São Paulo - SP  |  CEP: 03363-070",
                "Telefone: (11) 2784-4188  |  contato@pasquetti.com.br"]:
        ROW+=1; ws.row_dimensions[ROW].height = 13
        ws.merge_cells(f"C{ROW}:I{ROW}"); c = ws[f"C{ROW}"]
        c.value = txt; c.font = Font(name="Calibri",size=9,color="444444")
        c.alignment = Alignment(horizontal="right",vertical="center")

    # Linha cobre separadora
    ROW+=1; ws.row_dimensions[ROW].height = 4
    ws.merge_cells(f"A{ROW}:I{ROW}"); ws[f"A{ROW}"].fill = PatternFill("solid",fgColor=COR_COPPER)

    ROW+=1; ws.row_dimensions[ROW].height = 30
    ws.merge_cells(f"A{ROW}:E{ROW}"); c = ws[f"A{ROW}"]
    c.value = f"ORÇAMENTO Nº {num}"
    c.font = Font(name="Calibri",bold=True,size=17,color=COR_H)
    c.alignment = Alignment(horizontal="left",vertical="center")
    ws.merge_cells(f"F{ROW}:I{ROW}"); c = ws[f"F{ROW}"]
    c.value = datetime.date.today().strftime("%d/%m/%Y")
    c.font = Font(name="Calibri",size=11,color="666666")
    c.alignment = Alignment(horizontal="right",vertical="center")

    ROW+=1; ws.row_dimensions[ROW].height = 6

    # Cliente
    ROW+=1; ws.row_dimensions[ROW].height = 22
    ws.merge_cells(f"A{ROW}:I{ROW}"); c = ws[f"A{ROW}"]
    c.value = "Informações do Cliente"; c.border = bb()
    c.font = Font(name="Calibri",bold=True,size=12,color=COR_H)
    c.alignment = Alignment(horizontal="left",vertical="center")

    ROW+=1; ws.row_dimensions[ROW].height = 20
    ws.merge_cells(f"A{ROW}:I{ROW}"); c = ws[f"A{ROW}"]
    c.value = cliente.get("nome",""); c.font = Font(name="Calibri",bold=True,size=13)
    c.alignment = Alignment(horizontal="left",vertical="center")

    ROW+=1; ws.row_dimensions[ROW].height = 14
    ws.merge_cells(f"A{ROW}:D{ROW}"); ws[f"A{ROW}"].value = f"CNPJ: {cliente.get('cnpj','')}"
    ws[f"A{ROW}"].font = Font(name="Calibri",size=10)
    ws.merge_cells(f"E{ROW}:I{ROW}"); ws[f"E{ROW}"].value = cliente.get("endereco","")
    ws[f"E{ROW}"].font = Font(name="Calibri",size=10)

    ROW+=1; ws.row_dimensions[ROW].height = 14
    ws.merge_cells(f"A{ROW}:D{ROW}"); ws[f"A{ROW}"].value = f"Telefone: {cliente.get('telefone','')}"
    ws[f"A{ROW}"].font = Font(name="Calibri",bold=True,size=10)
    ws.merge_cells(f"E{ROW}:I{ROW}"); ws[f"E{ROW}"].value = f"Email: {cliente.get('email','')}"
    ws[f"E{ROW}"].font = Font(name="Calibri",bold=True,size=10)
    ROW+=1; ws.row_dimensions[ROW].height = 8

    # Tabela
    ROW+=1; ws.row_dimensions[ROW].height = 22
    ws.merge_cells(f"A{ROW}:I{ROW}"); c = ws[f"A{ROW}"]
    c.value = "Itens do ORÇAMENTO"; c.border = bb()
    c.font = Font(name="Calibri",bold=True,size=12,color=COR_H)
    c.alignment = Alignment(horizontal="left",vertical="center")

    ROW+=1; ws.row_dimensions[ROW].height = 22
    for j,(h,a) in enumerate(zip(
        ["Produto","NCM","Quant.","Unit. (R$)",_imp_lbl,"IPI","Unit. c/ Imp.","Valor Total (R$)"],
        ["left","center","center","right","right","right","right","right"]),2):
        c = ws.cell(row=ROW,column=j,value=h)
        c.font = Font(name="Calibri",bold=True,size=10,color="FFFFFF")
        c.fill = PatternFill("solid",fgColor=COR_H)
        c.alignment = Alignment(horizontal=a,vertical="center"); c.border = bd()

    # Linhas de itens — com FÓRMULAS (Unit c/ Imp e Valor Total recalculam sozinhos)
    first_data = ROW + 1
    for k,item in enumerate(confirmados):
        ROW+=1; ws.row_dimensions[ROW].height = 20
        bg = COR_L if k%2==0 else "FFFFFF"
        # B produto | C ncm | D qtd | E unit | F icms_st(unit) | G ipi(unit)
        # H unit c/ imp = E+F+G | I total = D*H
        valores = [item["produto"], item["ncm"], item["quantidade"], item["preco"],
                   float(item.get("st_unit", 0.0)), 0.0, f"=E{ROW}+F{ROW}+G{ROW}", f"=D{ROW}*H{ROW}"]
        fmts    = [None, None, '#,##0.###', '#,##0.0000', '#,##0.0000',
                   '#,##0.0000', '#,##0.0000', '#,##0.00']
        aligns  = ["left","center","center","right","right","right","right","right"]
        for j,(v,f,a) in enumerate(zip(valores, fmts, aligns), 2):
            c = ws.cell(row=ROW,column=j,value=v)
            c.font = Font(name="Calibri",size=9)
            c.fill = PatternFill("solid",fgColor=bg)
            c.alignment = Alignment(horizontal=a,vertical="center",wrap_text=(j==2))
            c.border = bd()
            if f: c.number_format = f
    has_rows = bool(confirmados)
    last_data = ROW if has_rows else None

    # Totais — fórmulas de soma
    if has_rows:
        rD, rE, rF = f"D{first_data}:D{last_data}", f"E{first_data}:E{last_data}", f"F{first_data}:F{last_data}"
        rG, rI = f"G{first_data}:G{last_data}", f"I{first_data}:I{last_data}"
        qtd_f      = f"=SUM({rD})"
        subtotal_f = f"=SUMPRODUCT({rD},{rE})"
        ipi_f      = f"=SUMPRODUCT({rD},{rG})"
        icms_f     = f"=SUMPRODUCT({rD},{rF})"
        total_f    = f"=SUM({rI})"
    else:
        qtd_f = subtotal_f = ipi_f = icms_f = total_f = 0

    ROW+=1; ws.row_dimensions[ROW].height = 6
    for lbl,val,bold,nf in [("Subtotal:",subtotal_f,False,'#,##0.00'),
                            ("IPI:",ipi_f,False,'#,##0.00'),
                            (f"{_imp_lbl}:",icms_f,False,'#,##0.00'),
                            ("Total:",total_f,True,'#,##0.00')]:
        ROW+=1; ws.row_dimensions[ROW].height = 18
        ws.merge_cells(f"B{ROW}:H{ROW}"); c = ws[f"B{ROW}"]
        c.value = lbl; c.font = Font(name="Calibri",bold=bold,size=10,color=COR_H if bold else "444444")
        c.alignment = Alignment(horizontal="right",vertical="center")
        c2 = ws.cell(row=ROW,column=9,value=val)
        c2.font = Font(name="Calibri",bold=bold,size=10,color=COR_H if bold else "444444")
        c2.alignment = Alignment(horizontal="right",vertical="center")
        c2.number_format = nf
        if bold: c2.border = bd()

    # Condições
    ROW+=2; ws.row_dimensions[ROW].height = 22
    ws.merge_cells(f"A{ROW}:I{ROW}"); c = ws[f"A{ROW}"]
    c.value = "Vencimentos e Condições"; c.border = bb()
    c.font = Font(name="Calibri",bold=True,size=12,color=COR_H)
    c.alignment = Alignment(horizontal="left",vertical="center")
    for lbl,val in [("Vendedor:",vendedor or "—"),
                    ("Pagamento:",condicao_pagamento or "À vista"),("Prazo da proposta:","10 dias"),
                    ("Obs.:","Preços sujeitos a alteração sem aviso prévio | Entrega sujeita a confirmação de estoque")]:
        ROW+=1; ws.row_dimensions[ROW].height = 16
        ws.cell(row=ROW,column=2,value=lbl).font = Font(name="Calibri",bold=True,size=10)
        ws.merge_cells(f"C{ROW}:I{ROW}")
        ws.cell(row=ROW,column=3,value=val).font = Font(name="Calibri",size=10)

    # Itens para revisão manual
    if sugestoes_manuais:
        ROW+=2; ws.row_dimensions[ROW].height = 20
        ws.merge_cells(f"A{ROW}:I{ROW}"); c = ws[f"A{ROW}"]
        c.value = "⚠ Itens com correspondência parcial (60–79%) — confirmar manualmente antes de incluir"
        c.font = Font(name="Calibri",bold=True,size=10,color="7a4f00")
        c.fill = PatternFill("solid",fgColor="FFF3CD")
        c.alignment = Alignment(horizontal="left",vertical="center")
        for s in sugestoes_manuais:
            ROW+=1; ws.row_dimensions[ROW].height = 16
            ws.cell(row=ROW,column=2,value=s["descricao"]).font = Font(name="Calibri",size=10,italic=True)
            ws.cell(row=ROW,column=4,value=f"Sugestão: {s['produto']}  ({s['score']:.0f}%)").font = Font(name="Calibri",size=10,color=COR_COPPER)
            ws.merge_cells(f"D{ROW}:I{ROW}")

    # Não encontrados
    if nao_enc:
        ROW+=2; ws.row_dimensions[ROW].height = 20
        ws.merge_cells(f"A{ROW}:I{ROW}"); c = ws[f"A{ROW}"]
        c.value = "✗ Itens NÃO ENCONTRADOS no catálogo — adicionar manualmente"
        c.font = Font(name="Calibri",bold=True,size=10,color="7a0000")
        c.fill = PatternFill("solid",fgColor="FFE0E0")
        c.alignment = Alignment(horizontal="left",vertical="center")
        for na in nao_enc:
            ROW+=1; ws.row_dimensions[ROW].height = 16
            ws.cell(row=ROW,column=2,value=f"{na['descricao']}  (qtd: {na['quantidade']:.0f})").font = Font(name="Calibri",size=10)

    # Rodapé
    ROW+=2; ws.row_dimensions[ROW].height = 14
    ws.merge_cells(f"A{ROW}:I{ROW}"); c = ws[f"A{ROW}"]
    c.value = f"Gerado em {datetime.datetime.now().strftime('%d/%m/%Y às %H:%M:%S')} | Pasquetti — Sistema de Cotações"
    c.font = Font(name="Calibri",size=8,color="999999",italic=True)
    c.alignment = Alignment(horizontal="right",vertical="center")

    # Aba de revisão
    ws2 = wb.create_sheet("Revisão Completa")
    for w,col in zip([40,42,10,12,12,10],range(1,7)):
        ws2.column_dimensions[get_column_letter(col)].width = w
    ws2.append(["Descrição Solicitada","Produto Correspondido","Qtd","Preço Unit.","Total","Status"])
    for it in confirmados: ws2.append([it["descricao"],it["produto"],it["quantidade"],it["preco"],it["total"],"CONFIRMADO"])
    for it in sugestoes_manuais: ws2.append([it["descricao"],it.get("produto",""),it["quantidade"],"","",f"SUGESTÃO {it['score']:.0f}%"])
    for it in nao_enc: ws2.append([it["descricao"],"NÃO ENCONTRADO",it["quantidade"],"","","NÃO ENCONTRADO"])

    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return buf.getvalue()


# ════════════════════════════════════════════════════════════════════════════
# INTERFACE
# ════════════════════════════════════════════════════════════════════════════

# ── Cabeçalho ──────────────────────────────────────────────────────────────
st.markdown(f"""
<div class="pq-header">
  <div>
    <p class="pq-title">Gerador de Cotações</p>
    <p class="pq-sub">Preencha os dados do cliente, insira os itens e clique em Gerar Cotação</p>
  </div>
  <span class="pq-tag">Pasquetti</span>
</div>
""", unsafe_allow_html=True)


# ── Sidebar ────────────────────────────────────────────────────────────────
with st.sidebar:
    _logo_v = logo_b64("vertical")
    if _logo_v:
        # Caixa branca para o logo (texto navy) aparecer no sidebar escuro.
        st.markdown(
            f'<div style="background:#fff;border-radius:12px;padding:14px;margin-bottom:16px;text-align:center;">'
            f'<img src="data:image/png;base64,{_logo_v}" style="max-width:100%;max-height:130px;object-fit:contain;" /></div>',
            unsafe_allow_html=True)
    else:
        st.markdown(f"""
        <div style="text-align:center;padding:16px 0;border-bottom:1px solid rgba(255,255,255,0.15);margin-bottom:12px;">
          <div style="display:inline-flex;align-items:center;gap:10px;">
            {LOGO_SVG}
            <div style="text-align:left;">
              <div style="color:white;font-size:16px;font-weight:700;letter-spacing:1px;">PASQUETTI</div>
              <div style="color:{COPPER};font-size:9px;font-weight:500;letter-spacing:.5px;">TUBOS E CONEXÕES</div>
            </div>
          </div>
        </div>
        """, unsafe_allow_html=True)

    st.markdown("### 👤 Dados do Cliente")

    for k in ["input_nome","input_fantasia","input_end","input_tel","input_email","input_cnpj"]:
        if k not in st.session_state: st.session_state[k] = ""

    # Busca ÚNICA do cliente: procura na base e, se não achar, oferece a SEFAZ no MESMO campo
    if dados_supabase and dados_supabase.disponivel() and st_searchbox is not None:
        def _busca_cli(q):
            out = []
            _regs = dados_supabase.buscar_clientes(q)
            for c in _regs:
                out.append((f"💾 {_fmt_cnpj(c.get('cnpj',''))} — {c.get('nome','')}", "BASE::"+c.get("cnpj","")))
            _dig = "".join(ch for ch in (q or "") if ch.isdigit())
            if len(_dig) >= 8 and not any(c.get("cnpj")==_dig for c in _regs):
                _rot = _fmt_cnpj(_dig) if len(_dig)==14 else _dig
                out.append((f"🌐 Buscar {_rot} na SEFAZ", "SEFAZ::"+_dig))
            return out
        _sel = st_searchbox(_busca_cli, key="cli_search",
                            placeholder="🔎 cliente: digite CNPJ ou nome")
        if _sel and st.session_state.get("cli_carregado") != _sel:
            st.session_state["cli_carregado"] = _sel
            if _sel.startswith("BASE::"):
                _cn = _sel[6:]
                _rs = dados_supabase.buscar_clientes(_cn)
                _r = next((x for x in _rs if x.get("cnpj")==_cn), (_rs[0] if _rs else None))
                if _r:
                    _fill_cliente(_r)
                    flash(f"Cliente “{_r.get('nome','')[:40]}” carregado do cadastro.", "success")
                    st.rerun()
            elif _sel.startswith("SEFAZ::"):
                _cn = _sel[7:]
                with st.spinner("Consultando SEFAZ..."):
                    _d, _e = buscar_cnpj(_cn)
                if _d:
                    _fill_cliente({"cnpj":_cn,"nome":_d["nome"],"fantasia":_d.get("fantasia",""),
                                   "endereco":_d["endereco"],"telefone":_d["telefone"],
                                   "email":_d["email"],"uf":_d.get("uf","")})
                    _salvo = False
                    try:
                        dados_supabase.salvar_cliente(_cn,_d["nome"],_d.get("fantasia",""),_d["endereco"],
                                                      _d["telefone"],_d["email"],_d.get("uf",""))
                        _salvo = True
                    except Exception:
                        pass
                    flash(f"Cliente “{_d['nome'][:40]}” encontrado na SEFAZ"
                          + (" e salvo no cadastro." if _salvo else "."), "success")
                    st.rerun()
                else:
                    st.error(f"❌ {_e}")
        cnpj_input = st.session_state.get("input_cnpj","")
        if cnpj_input:
            st.caption(f"📋 CNPJ: **{cnpj_input}**")
    else:
        # Modo simples (sem o componente de busca): o PRÓPRIO campo CNPJ faz a busca base/SEFAZ
        cnpj_input = st.text_input("CNPJ (opcional)", key="input_cnpj", placeholder="00.000.000/0001-00",
                                   help="Opcional. Digite o CNPJ e clique em Buscar — procura na base e, se não achar, na SEFAZ. "
                                        "A cotação pode ser gerada sem CNPJ.")
        if st.button("🔎 Buscar cliente (base / SEFAZ)", use_container_width=True) and cnpj_input.strip():
            _rs = dados_supabase.buscar_clientes(cnpj_input) if (dados_supabase and dados_supabase.disponivel()) else []
            _dig = "".join(ch for ch in cnpj_input if ch.isdigit())
            _r = next((x for x in _rs if x.get("cnpj")==_dig), (_rs[0] if _rs else None))
            if _r:
                st.session_state["input_nome"]=_r.get("nome","") or ""
                st.session_state["input_fantasia"]=_r.get("fantasia","") or ""
                st.session_state["input_end"]=_r.get("endereco","") or ""
                st.session_state["input_tel"]=_r.get("telefone","") or ""
                st.session_state["input_email"]=_r.get("email","") or ""
                if _r.get("uf"): st.session_state["uf_sel"]=_r["uf"]
                flash(f"Cliente “{(_r.get('nome','') or '')[:40]}” carregado do cadastro.", "success")
                st.rerun()
            else:
                with st.spinner("Consultando SEFAZ..."):
                    _d, _e = buscar_cnpj(cnpj_input)
                if _d:
                    st.session_state["input_nome"]=_d["nome"]; st.session_state["input_fantasia"]=_d.get("fantasia","")
                    st.session_state["input_end"]=_d["endereco"]; st.session_state["input_tel"]=_d["telefone"]
                    st.session_state["input_email"]=_d["email"]
                    if _d.get("uf"): st.session_state["uf_sel"]=_d["uf"]
                    _salvo = False
                    if dados_supabase and dados_supabase.disponivel():
                        try:
                            dados_supabase.salvar_cliente(_dig,_d["nome"],_d.get("fantasia",""),_d["endereco"],
                                                          _d["telefone"],_d["email"],_d.get("uf",""))
                            _salvo = True
                        except Exception:
                            pass
                    flash(f"Cliente “{_d['nome'][:40]}” encontrado na SEFAZ"
                          + (" e salvo no cadastro." if _salvo else "."), "success")
                    st.rerun()
                else:
                    st.error(f"❌ {_e}")
    nome     = st.text_input("Nome / Razão Social (opcional)", placeholder="Construtora ABC Ltda", key="input_nome")
    fantasia = st.text_input("Nome Fantasia", placeholder="(opcional)", key="input_fantasia")
    endereco = st.text_input("Endereço", placeholder="Rua X, 123 - Bairro - Cidade", key="input_end")
    telefone = st.text_input("Telefone", placeholder="(11) 9999-9999", key="input_tel")
    email    = st.text_input("E-mail", placeholder="contato@cliente.com.br", key="input_email")

    if dados_supabase and dados_supabase.disponivel():
        if st.button("💾 Salvar cliente no cadastro", use_container_width=True, key="salvar_cli"):
            if (cnpj_input or "").strip() and (nome or "").strip():
                try:
                    dados_supabase.salvar_cliente(cnpj_input, nome, fantasia, endereco, telefone, email,
                                                  st.session_state.get("uf_sel",""))
                    st.success("✅ Cliente salvo no cadastro.")
                except Exception as e:
                    st.error(f"Erro ao salvar: {e}")
            else:
                st.warning("Preencha ao menos CNPJ e Nome para salvar.")

    st.markdown("---")
    st.markdown("### ⚙️ Configurações")
    vendedor = st.selectbox("Vendedor", options=["Marcelo","Guilherme"])
    _UFS = ["AC","AL","AP","AM","BA","CE","DF","ES","GO","MA","MT","MS","MG","PA",
            "PB","PR","PE","PI","RJ","RN","RS","RO","RR","SC","SP","SE","TO"]
    if "uf_sel" not in st.session_state: st.session_state["uf_sel"] = "SP"
    uf_destino = st.selectbox("UF de destino (para ST)", options=_UFS, key="uf_sel",
                              help="Usada para aplicar a alíquota de ST por NCM. Preenchida pela busca de CNPJ.")
    tipo_cliente = st.selectbox(
        "Tipo de cliente (ICMS)",
        options=["Revenda / contribuinte (ST)", "Consumidor final / não-contribuinte (DIFAL)"],
        key="tipo_cliente",
        help="Revenda → calcula ICMS-ST (com MVA). Consumidor final → calcula DIFAL (diferencial de alíquota).")
    consumidor_final = tipo_cliente.startswith("Consumidor")
    tabela = st.selectbox("Tabela de preços", options=["consumo","revenda","pressao","smu","todos"],
        format_func=lambda x: {"consumo":"🏠 Consumo — HL Mar/2026",
                                "revenda":"🏪 Revenda — HL Mar/2026",
                                "pressao":"💧 Pressão — JGS Abr/2026",
                                "smu":    "🔩 SMU — Fev/2026",
                                "todos":  "📋 Todas as tabelas"}[x])
    # ── Condição de pagamento ──
    BASE_COND = ["À vista","Boleto 28 dias","30/60/90 dias","28/42/56 dias","45/60 dias"]
    if dados_supabase and dados_supabase.disponivel():
        cond_lista = dados_supabase.listar_condicoes_pagamento() or BASE_COND
    else:
        cond_lista = BASE_COND
    if "cond_pagamento" not in st.session_state or st.session_state["cond_pagamento"] not in cond_lista:
        st.session_state["cond_pagamento"] = cond_lista[0]
    _OUTRA = "✏️ Outra (digitar)…"
    _opts = cond_lista + [_OUTRA]
    _cur = st.session_state["cond_pagamento"]
    cond_sel = st.selectbox("Condição de pagamento", _opts,
                            index=_opts.index(_cur) if _cur in _opts else 0, key="cond_select")
    if cond_sel == _OUTRA:
        nova = st.text_input("Nova condição", key="cond_nova_txt", placeholder="ex.: 30/60 dias")
        if nova.strip():
            sim = cond_similar(nova, cond_lista)
            if sim and normalizar(sim) != normalizar(nova):
                st.info(f"Parecida com já cadastrada: **{sim}**")
                _b1, _b2 = st.columns(2)
                if _b1.button(f"Usar “{sim}”", key="cond_usar", use_container_width=True):
                    st.session_state["cond_pagamento"] = sim
                    flash(f"Condição de pagamento “{sim}” selecionada.", "info")
                    st.rerun()
                if _b2.button("Criar nova", key="cond_criar", use_container_width=True):
                    _ok = True
                    if dados_supabase and dados_supabase.disponivel():
                        try: dados_supabase.salvar_condicao_pagamento(nova)
                        except Exception as e:
                            _ok = False; st.error(f"Erro ao salvar: {e}")
                    st.session_state["cond_pagamento"] = nova
                    if _ok: flash(f"Condição de pagamento “{nova}” cadastrada e selecionada.", "success")
                    st.rerun()
            else:
                if st.button("➕ Cadastrar e usar", key="cond_add", use_container_width=True):
                    _ok = True
                    if dados_supabase and dados_supabase.disponivel():
                        try: dados_supabase.salvar_condicao_pagamento(nova)
                        except Exception as e:
                            _ok = False; st.error(f"Erro ao salvar: {e}")
                    st.session_state["cond_pagamento"] = nova
                    if _ok: flash(f"Condição de pagamento “{nova}” cadastrada e selecionada.", "success")
                    st.rerun()
            cond_pagamento = nova
        else:
            cond_pagamento = st.session_state["cond_pagamento"]
    else:
        st.session_state["cond_pagamento"] = cond_sel
        cond_pagamento = cond_sel
    st.caption(f"Na cotação: **{cond_pagamento}**")

    num_orcamento = st.text_input("Nº do Orçamento", placeholder="Gerado automaticamente")
    if not num_orcamento:
        import random; num_orcamento = str(random.randint(1000,9999))

    st.markdown("---")
    if not CATALOG_JSON.exists():
        st.error("⚠️ Catálogo não encontrado!\nRode extrair_catalogo.py primeiro.")
    else:
        with open(CATALOG_JSON) as f: cat_data = json.load(f)
        st.markdown(f"""<div style="background:rgba(255,255,255,0.08);border-radius:6px;padding:8px 10px;font-size:11px;color:#90c090;">
        ✅ <strong style="color:white">{len(cat_data)} produtos</strong> no catálogo<br>
        <span style="color:#aaa">Precisão mínima: 80%</span>
        </div>""", unsafe_allow_html=True)

    # ── Interpretação com IA ──────────────────────────────────────────────
    st.markdown("---")
    st.markdown("### 🤖 Interpretação com IA")
    usar_ia = False
    if matcher_ia is None:
        st.caption("Biblioteca 'anthropic' não instalada. Rode o instalador novamente.")
    else:
        cfg_ia = matcher_ia.carregar_config()
        tem_chave = bool(cfg_ia["api_key"])
        if tem_chave:
            usar_ia = st.toggle("Usar IA para interpretar itens", value=True,
                                help="Entende pedidos vagos/abreviados e lê foto/PDF. Custa ~centavos por cotação.")
            mascara = cfg_ia["api_key"][:7] + "…" + cfg_ia["api_key"][-4:]
            st.caption(f"🔑 Chave configurada ({mascara}) · modelo: {cfg_ia['modelo']}")
            with st.expander("Trocar chave / modelo"):
                nova = st.text_input("Chave de API (sk-ant-...)", type="password", key="ia_key_new")
                modelo = st.text_input("Modelo", value=cfg_ia["modelo"], key="ia_model_new")
                if st.button("Salvar configuração de IA"):
                    matcher_ia.salvar_config(nova or cfg_ia["api_key"], modelo)
                    st.success("Configuração salva. Recarregue a página.")
        else:
            st.caption("Cole sua chave de API da Anthropic para ligar a interpretação por IA.")
            nova = st.text_input("Chave de API (sk-ant-...)", type="password", key="ia_key")
            modelo = st.text_input("Modelo", value=matcher_ia.MODELO_PADRAO, key="ia_model")
            if st.button("Salvar e ativar IA") and nova.strip():
                matcher_ia.salvar_config(nova, modelo)
                st.success("Chave salva. Recarregue a página para ativar.")

    # ── Substituição Tributária (ST) ──────────────────────────────────────
    st.markdown("---")
    st.markdown("### 📐 Regras de ST")
    if dados_supabase is None or not dados_supabase.disponivel():
        st.caption("Banco de aprendizado não configurado — ST não será aplicada.")
    else:
        st.caption("Alíquota de ST por NCM e UF de destino. O sistema reaplica sozinho nas próximas cotações.")
        with st.expander("➕ Cadastrar / atualizar regra de ST"):
            st_ncm = st.text_input("NCM", placeholder="7307.11.00", key="st_ncm")
            cuf, cal = st.columns(2)
            st_uf  = cuf.selectbox("UF", options=_UFS, index=_UFS.index(st.session_state.get("uf_sel","SP")), key="st_uf")
            st_aliq = cal.number_input("Alíquota ST (%)", min_value=0.0, max_value=100.0, step=0.5, key="st_aliq")
            if st.button("Salvar regra de ST"):
                if st_ncm.strip():
                    try:
                        dados_supabase.salvar_regra_st(st_ncm, st_uf, st_aliq)
                        st.success(f"Regra salva: NCM {st_ncm} / {st_uf} = {st_aliq:.1f}%")
                    except Exception as e:
                        st.error(f"Erro ao salvar: {e}")
                else:
                    st.warning("Informe o NCM.")


# ── Área principal ──────────────────────────────────────────────────────────
tab1, tab2, tab3 = st.tabs(["📤  Carregar Arquivo", "✏️  Digitar / Colar", "📊  Tabela Interativa"])
itens_brutos = []

# Tab 1: Upload
with tab1:
    st.markdown('<span class="pq-section">Envie o arquivo da solicitação</span>', unsafe_allow_html=True)
    st.caption("Formatos aceitos: `.txt` · `.xlsx` · `.pdf` · `.jpg` · `.png`  (foto e PDF usam a IA)")
    arquivo = st.file_uploader("Arquivo", type=["txt","xlsx","xls","pdf","jpg","jpeg","png","bmp"],
                                label_visibility="collapsed")
    if arquivo:
        ext = Path(arquivo.name).suffix.lower()
        data = arquivo.read()
        ia_ok = matcher_ia is not None and matcher_ia.ia_disponivel()
        if ext == ".txt":
            itens_brutos = parse_texto(data.decode("utf-8", errors="ignore"))
            st.success(f"✅ **{len(itens_brutos)} itens** lidos do arquivo de texto")
        elif ext in (".xlsx",".xls"):
            try:
                itens_brutos = parse_xlsx_bytes(data)
            except ValueError as e:
                st.error(f"❌ {e}"); itens_brutos = []
            except Exception as e:
                st.error(f"❌ Erro ao ler a planilha: {e}"); itens_brutos = []
            if itens_brutos:
                st.success(f"✅ **{len(itens_brutos)} itens** lidos da planilha")
            elif data:
                st.warning("Li a planilha mas não encontrei itens. Confira se há uma coluna de "
                           "**Produto/Descrição** (e, opcional, **Quantidade**) com dados preenchidos. "
                           "Se as células forem fórmulas, abra e salve a planilha no Excel antes de enviar.")
        elif ext == ".pdf":
            if ia_ok:
                with st.spinner("Lendo o PDF com a IA..."):
                    try:
                        itens_brutos = matcher_ia.extrair_itens_de_pdf(data)
                    except Exception as e:
                        st.error(f"Erro ao ler o PDF com a IA: {e}"); itens_brutos = []
                if itens_brutos: st.success(f"✅ **{len(itens_brutos)} itens** lidos do PDF")
                else: st.warning("Não consegui extrair itens do PDF.")
            else:
                st.warning("Leitura de PDF requer a IA ligada (configure a chave na barra lateral).")
        elif ext in (".jpg",".jpeg",".png",".bmp"):
            mt = {"jpg":"image/jpeg","jpeg":"image/jpeg","png":"image/png","bmp":"image/bmp"}[ext.strip(".")]
            if ia_ok:
                with st.spinner("Lendo a imagem com a IA..."):
                    try:
                        itens_brutos = matcher_ia.extrair_itens_de_imagem(data, media_type=mt)
                    except Exception as e:
                        st.error(f"Erro ao ler a imagem com a IA: {e}"); itens_brutos = []
                if itens_brutos: st.success(f"✅ **{len(itens_brutos)} itens** lidos da imagem")
                else: st.warning("Não consegui extrair itens da imagem.")
            else:
                with st.spinner("Reconhecendo texto na imagem (OCR)..."):
                    itens_brutos = parse_imagem_bytes(data)
                if itens_brutos: st.success(f"✅ **{len(itens_brutos)} itens** reconhecidos (OCR)")
                else: st.warning("Não foi possível reconhecer itens. Ligue a IA ou use o modo de digitação.")
        if itens_brutos:
            with st.expander("👁  Ver itens lidos"):
                for it in itens_brutos:
                    st.write(f"• **{it['quantidade']:.0f}x** {it['descricao']}")

# Tab 2: Texto livre
with tab2:
    st.markdown('<span class="pq-section">Cole ou digite os itens</span>', unsafe_allow_html=True)
    st.caption("Um item por linha — coloque a **quantidade no início**:")
    texto_livre = st.text_area("Itens", height=270, label_visibility="collapsed",
        placeholder="90  joelho fofo 100mm\n72  JF 75mm 90 graus\n6   anel 150mm\n190 anel borracha 100mm")
    if texto_livre.strip():
        itens_brutos = parse_texto(texto_livre)
        if itens_brutos: st.success(f"✅ **{len(itens_brutos)} itens** prontos")

# Tab 3: Tabela editável
with tab3:
    import pandas as pd
    st.markdown('<span class="pq-section">Preencha a tabela de itens</span>', unsafe_allow_html=True)
    if "tabela_itens" not in st.session_state:
        st.session_state.tabela_itens = pd.DataFrame({"Produto / Descrição":[""] * 8, "Quantidade":[1]*8})

    c1, c2, _ = st.columns([1,1,4])
    with c1:
        if st.button("➕ Adicionar linhas"):
            novas = pd.DataFrame({"Produto / Descrição":[""]*5,"Quantidade":[1]*5})
            st.session_state.tabela_itens = pd.concat([st.session_state.tabela_itens,novas],ignore_index=True)
    with c2:
        if st.button("🗑 Limpar"):
            st.session_state.tabela_itens = pd.DataFrame({"Produto / Descrição":[""] * 8,"Quantidade":[1]*8})

    tabela_edit = st.data_editor(
        st.session_state.tabela_itens, use_container_width=True, num_rows="dynamic",
        column_config={
            "Produto / Descrição": st.column_config.TextColumn("Produto / Descrição",
                help="Ex: joelho fofo 90 150mm | JF 100mm | anel borracha 75mm", width="large"),
            "Quantidade": st.column_config.NumberColumn("Qtd", min_value=0, step=1, width="small"),
        }, hide_index=True, key="editor_tabela")
    st.session_state.tabela_itens = tabela_edit

    itens_tab = []
    for _, row in tabela_edit.iterrows():
        desc = str(row.get("Produto / Descrição","")).strip()
        try: qtd = float(row.get("Quantidade",1) or 1)
        except: qtd = 1
        if desc and desc.lower() not in ("none","nan",""): itens_tab.append({"descricao":desc,"quantidade":qtd})
    if itens_tab:
        itens_brutos = itens_tab
        st.success(f"✅ **{len(itens_tab)} itens** preenchidos")


# ── Botão Gerar ────────────────────────────────────────────────────────────
st.markdown("---")
col_btn, col_hint = st.columns([2,3])
with col_btn:
    gerar = st.button("🧾 Gerar Cotação", type="primary",
                      disabled=(not itens_brutos or not CATALOG_JSON.exists()))
with col_hint:
    if not itens_brutos:
        st.info("Insira os itens em uma das abas acima.")
    elif not nome:
        st.info("💡 Cliente não preenchido — a cotação será gerada normalmente para conferência. "
                "Preencher nome/CNPJ é opcional.")


# ── Resultado ──────────────────────────────────────────────────────────────
if gerar and itens_brutos:
    spin_msg = "Interpretando itens com IA e buscando no catálogo..." if usar_ia \
               else "Buscando correspondências no catálogo (precisão ≥ 80%)..."
    with st.spinner(spin_msg):
        catalogo  = carregar_catalogo(tabela)
        _disp     = bool(dados_supabase and dados_supabase.disponivel())
        correcoes = dados_supabase.listar_correcoes(tabela) if _disp else {}
        nao_trab_set = dados_supabase.listar_nao_trabalhados() if _disp else set()
        conf, sug, nao, nao_trab = processar_hibrido(itens_brutos, catalogo, usar_ia, correcoes, nao_trab_set)

    # Aplicar ICMS-ST (revenda) ou DIFAL (consumidor final) por item, por NCM + UF
    regras_st = dados_supabase.listar_regras_st() if (dados_supabase and dados_supabase.disponivel()) else {}
    for it in conf:
        if dados_supabase:
            it["st_unit"] = dados_supabase.calcular_st_difal(
                it["preco"], it.get("ncm", ""), uf_destino,
                regras=regras_st, consumidor_final=consumidor_final)
        else:
            it["st_unit"] = 0.0

    subtotal = sum(i["total"] for i in conf)
    total_st = sum(i.get("st_unit",0.0) * i["quantidade"] for i in conf)
    st.session_state["cotacao"] = {"conf":conf,"sug":sug,"nao":nao,"nao_trab":nao_trab,
        "subtotal":subtotal,"total_st":total_st,"tabela":tabela,"num_orcamento":num_orcamento,
        "itens_brutos":itens_brutos,"usar_ia":usar_ia,"consumidor_final":consumidor_final}
    # Cadastra/atualiza o cliente automaticamente para próximas cotações
    if dados_supabase and dados_supabase.disponivel() and nome and cnpj_input:
        try:
            dados_supabase.salvar_cliente(cnpj_input, nome, fantasia, endereco, telefone, email,
                                          st.session_state.get("uf_sel",""))
        except Exception:
            pass

# Render a partir do session_state (sobrevive ao rerun do botão de ensinar/salvar)
if st.session_state.get("cotacao"):
    _C = st.session_state["cotacao"]
    conf, sug, nao = _C["conf"], _C["sug"], _C["nao"]
    nao_trab = _C.get("nao_trab", [])
    subtotal, total_st = _C["subtotal"], _C["total_st"]
    tabela = _C.get("tabela", tabela)
    num_orcamento = _C.get("num_orcamento", num_orcamento)

    # Métricas
    st.markdown("---")
    st.markdown('<span class="pq-section">Resultado da Cotação</span>', unsafe_allow_html=True)

    m1,m2,m3,m4,m5 = st.columns(5)
    with m1: st.markdown(f'<div class="pq-metric green"><div class="pq-metric-val green">{len(conf)}</div><div class="pq-metric-lbl">Confirmados (≥80%)</div></div>', unsafe_allow_html=True)
    with m2: st.markdown(f'<div class="pq-metric copper"><div class="pq-metric-val copper">{len(sug)}</div><div class="pq-metric-lbl">Sugestões (60–79%)</div></div>', unsafe_allow_html=True)
    with m3: st.markdown(f'<div class="pq-metric red"><div class="pq-metric-val red">{len(nao)}</div><div class="pq-metric-lbl">Não encontrados</div></div>', unsafe_allow_html=True)
    with m4: st.markdown(f'<div class="pq-metric"><div class="pq-metric-val">{len(nao_trab)}</div><div class="pq-metric-lbl">Não trabalhamos</div></div>', unsafe_allow_html=True)
    with m5: st.markdown(f'<div class="pq-metric"><div class="pq-metric-val">R$ {subtotal:,.2f}</div><div class="pq-metric-lbl">Subtotal (confirmados)</div></div>', unsafe_allow_html=True)

    # Tabela de confirmados
    if conf:
        st.markdown('<span class="pq-section" style="color:#27ae60">✅ Itens Confirmados — entram na cotação</span>', unsafe_allow_html=True)
        linhas = ""
        for it in conf:
            s = it["score"]
            badge = f'<span class="badge badge-a">{s:.0f}%</span>' if s>=90 else f'<span class="badge badge-m">{s:.0f}%</span>'
            linhas += f"""<tr>
              <td style="color:#555;font-style:italic;font-size:11px">{it['descricao']}</td>
              <td><strong>{it['produto']}</strong></td>
              <td style="text-align:center">{it['quantidade']:.0f}</td>
              <td style="text-align:right">R$ {it['preco']:,.4f}</td>
              <td style="text-align:right"><strong>R$ {it['total']:,.2f}</strong></td>
              <td style="text-align:center">{badge}</td>
            </tr>"""
        st.markdown(f"""<table class="pq-table">
          <thead><tr>
            <th>Solicitado</th><th>Produto no Catálogo</th>
            <th style="text-align:center">Qtd</th>
            <th style="text-align:right">Preço Unit.</th>
            <th style="text-align:right">Total</th>
            <th style="text-align:center">Score</th>
          </tr></thead>
          <tbody>{linhas}</tbody>
        </table>
        <p style="text-align:right;font-weight:700;font-size:15px;color:{NAVY};margin-top:10px;">
          Subtotal: R$ {subtotal:,.2f}
        </p>""", unsafe_allow_html=True)

    # Sugestões (não entram automaticamente)
    if sug:
        sug_rows = ""
        for s in sug:
            just = s.get("justificativa", "")
            just_html = f'<br><span style="color:#999;font-size:10px">{just}</span>' if just else ""
            sug_rows += f"""<tr>
              <td style="color:#555;font-style:italic">{s['descricao']}</td>
              <td>{s.get('produto','—')}{just_html}</td>
              <td style="text-align:center">{s['quantidade']:.0f}</td>
              <td style="text-align:center"><span class="badge badge-s">{s['score']:.0f}%</span></td>
            </tr>"""
        st.markdown(f"""<div class="pq-suggest">
          <div class="pq-suggest-title">⚠️ Sugestões com correspondência parcial (60–79%) — NÃO incluídas na cotação</div>
          <p style="color:#7a5a20;margin-bottom:10px;font-size:11px;">
            Revise cada item abaixo. Se a sugestão estiver correta, confirme e inclua manualmente na cotação.
          </p>
          <table class="pq-table" style="font-size:11px;">
            <thead><tr>
              <th>Solicitado</th><th>Melhor Sugestão</th>
              <th style="text-align:center">Qtd</th>
              <th style="text-align:center">Precisão</th>
            </tr></thead>
            <tbody>{sug_rows}</tbody>
          </table>
        </div>""", unsafe_allow_html=True)

    # Não encontrados (com opção de marcar "não trabalhamos")
    if nao:
        st.markdown('<div class="pq-notfound"><div class="pq-notfound-title">'
                    '❌ Itens não encontrados no catálogo</div></div>', unsafe_allow_html=True)
        _disp_nt = bool(dados_supabase and dados_supabase.disponivel())
        for j, n in enumerate(nao):
            c1, c2 = st.columns([4, 2])
            with c1:
                st.markdown(f"<div style='padding-top:6px;font-size:13px'>• <strong>{n['descricao']}</strong> "
                            f"<span style='color:#888'>(qtd: {n['quantidade']:.0f})</span></div>",
                            unsafe_allow_html=True)
            with c2:
                if _disp_nt and st.button("🚫 Não trabalhamos com este item",
                                          key=f"nt_{num_orcamento}_{j}", use_container_width=True):
                    try:
                        dados_supabase.salvar_nao_trabalhado(n["descricao"], vendedor)
                        flash(f"“{n['descricao'][:40]}” marcado como não trabalhamos — fora da cotação.", "success")
                    except Exception as e:
                        st.error(f"Erro ao salvar: {e}")
                    st.session_state["cotacao"]["nao"] = [x for k, x in enumerate(nao) if k != j]
                    st.session_state["cotacao"]["nao_trab"] = list(nao_trab) + [
                        {"descricao": n["descricao"], "quantidade": n["quantidade"],
                         "conf": "NÃO TRABALHAMOS"}]
                    st.rerun()
        if _disp_nt:
            st.caption("Marque os que vocês não fornecem — eles saem da cotação e, "
                       "nas próximas vezes, já vêm classificados sozinhos.")

    # Itens que não trabalhamos (não entram no Excel enviado ao cliente)
    if nao_trab:
        nt_rows = "".join(f"<li><strong>{n['descricao']}</strong> (qtd: {n['quantidade']:.0f})</li>"
                          for n in nao_trab)
        st.markdown(f"""<div style="background:#f1f1f4;border:1px solid #d6d6dd;border-radius:10px;
              padding:12px 16px;margin-top:10px;">
          <div style="font-weight:700;color:#555;font-size:13px;">🚫 Itens que não trabalhamos — fora da cotação</div>
          <ul style="margin:6px 0;padding-left:18px;font-size:12px;color:#666;">{nt_rows}</ul>
          <div style="font-size:11px;color:#999;">Não aparecem no Excel enviado ao cliente.</div>
        </div>""", unsafe_allow_html=True)
        if dados_supabase and dados_supabase.disponivel():
            for j, n in enumerate(nao_trab):
                if st.button(f"↩︎ Desfazer “{n['descricao'][:40]}”",
                             key=f"undo_nt_{num_orcamento}_{j}"):
                    try:
                        dados_supabase.remover_nao_trabalhado(n["descricao"])
                        flash(f"“{n['descricao'][:40]}” voltou a ser cotado normalmente.", "info")
                    except Exception as e:
                        st.error(f"Erro: {e}")
                    st.session_state["cotacao"]["nao_trab"] = [
                        x for k, x in enumerate(nao_trab) if k != j]
                    st.rerun()

    # Download
    st.markdown("---")
    tabelas_nome = {"consumo":"TABELA HL CONSUMO — Março/2026","revenda":"TABELA HL REVENDA — Março/2026",
                    "pressao":"TABELA PRESSÃO JGS — Abril/2026","smu":"TABELA SMU — Fevereiro/2026","todos":"MÚLTIPLAS TABELAS"}
    cliente_dict = {"nome":nome,"cnpj":cnpj_input,"endereco":endereco,"telefone":telefone,"email":email}
    xlsx_bytes = gerar_xlsx_bytes(conf, sug, nao, num_orcamento, cliente_dict, tabelas_nome[tabela], vendedor, condicao_pagamento=cond_pagamento, consumidor_final=_C.get("consumidor_final", False))
    nome_arq = f"cotacao_{num_orcamento}_{nome.replace(' ','_')[:18]}.xlsx" if nome else f"cotacao_{num_orcamento}.xlsx"

    dcol, icol = st.columns([1,2])
    with dcol:
        st.download_button("⬇️  Baixar Cotação (.xlsx)", data=xlsx_bytes, file_name=nome_arq,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True)
    with icol:
        _imp_nome = "DIFAL" if _C.get("consumidor_final", False) else "ST"
        extra_st = f" · {_imp_nome}: R$ {total_st:,.2f}" if total_st else ""
        st.info(f"📄 Orçamento **Nº {num_orcamento}** · {len(conf)} itens confirmados · R$ {subtotal:,.2f}{extra_st}")

    # ── Revisar, marcar OK e corrigir — tabela editável ───────────────────────
    st.markdown("---")
    st.markdown('<span class="pq-section">🧠 Revisar itens — marque OK ou corrija o produto</span>', unsafe_allow_html=True)
    _disp_corr = bool(dados_supabase and dados_supabase.disponivel())
    if not _disp_corr:
        st.warning("Sem conexão com o banco de aprendizado agora — você pode revisar e corrigir, "
                   "mas as correções só serão memorizadas quando reconectar.")
    st.caption("Marque ✔ OK nos itens certos (marcar um item que não veio confirmado também o ensina). "
               "Para corrigir, digite termos do produto em qualquer ordem (ex.: \"ralo smu\") na lista suspensa.")
    catalogo = carregar_catalogo(tabela)
    _norm_base = [(p["descricao"], normalizar(p["descricao"])) for p in catalogo]
    revisaveis = conf + sug + nao
    _icone = {"CONFIRMADO":"✅","SUGESTÃO":"⚠️","NÃO ENCONTRADO":"❌"}

    hc = st.columns([3, 4, 1, 1])
    for _c, _t in zip(hc, ["Solicitado", "Produto no catálogo (digite p/ buscar)", "Qtd", "OK"]):
        _c.markdown(f"<div style='font-weight:700;color:{NAVY};font-size:12px'>{_t}</div>", unsafe_allow_html=True)

    escolhas = {}
    for idx, it in enumerate(revisaveis):
        atual = it.get("produto", "")
        ic = _icone.get(it.get("conf",""), "")
        r1, r2, r3, r4 = st.columns([3, 4, 1, 1])
        with r1:
            st.markdown(f"<div style='font-size:12px;padding-top:6px'>{ic} <strong>{it['descricao']}</strong></div>",
                        unsafe_allow_html=True)
        with r2:
            if st_searchbox is not None:
                def _busca(q, _base=_norm_base, _atual=atual):
                    if not q or not q.strip():
                        return [_atual] if _atual else []
                    toks = normalizar(q).split()
                    return [d for d, dn in _base if all(t in dn for t in toks)][:25]
                sel = st_searchbox(_busca, key=f"corr_{num_orcamento}_{idx}",
                                   placeholder="digite termos — ex.: ralo smu", default=atual or None)
                escolhas[idx] = sel or atual or "— manter —"
            else:
                q = st.text_input("p", key=f"busca_{num_orcamento}_{idx}",
                                  placeholder="digite termos — ex.: ralo smu", label_visibility="collapsed")
                if q.strip():
                    toks = normalizar(q).split()
                    matches = [d for d, dn in _norm_base if all(t in dn for t in toks)][:10]
                else:
                    matches = [atual] if atual else []
                _opc = ([atual] if atual else ["— manter —"]) + [m for m in matches if m != atual]
                escolhas[idx] = st.selectbox("p", _opc, key=f"selc_{num_orcamento}_{idx}",
                                             label_visibility="collapsed")
        with r3:
            st.markdown(f"<div style='font-size:12px;padding-top:6px;text-align:center'>{it['quantidade']:.0f}</div>",
                        unsafe_allow_html=True)
        with r4:
            st.checkbox("ok", key=f"ok_{num_orcamento}_{idx}",
                        value=(it.get("conf") == "CONFIRMADO"), label_visibility="collapsed")

    if st.button("💾 Salvar revisão e refazer cotação", key=f"salvar_{num_orcamento}", type="primary"):
        if not _disp_corr:
            st.warning("Sem conexão com o banco — não foi possível memorizar as correções agora.")
        else:
            n_salvos = 0
            _desc_set = {p["descricao"] for p in catalogo}
            for idx, it in enumerate(revisaveis):
                escolhido = escolhas.get(idx)
                orig = it.get("produto", "")
                final = escolhido if (escolhido and escolhido != "— manter —") else orig
                ok = bool(st.session_state.get(f"ok_{num_orcamento}_{idx}", False))
                if not final or final not in _desc_set:
                    continue
                muda = final != orig
                confirma = ok and it.get("conf") != "CONFIRMADO"
                if muda or confirma:
                    try:
                        dados_supabase.salvar_correcao(it["descricao"], final, tabela, vendedor)
                        n_salvos += 1
                    except Exception as e:
                        st.error(f"Erro ao salvar '{it['descricao']}': {e}")
            if n_salvos:
                # Refaz a cotação automaticamente já com as correções aprendidas
                _ib = _C.get("itens_brutos", [])
                _ia = _C.get("usar_ia", False)
                _cat = carregar_catalogo(tabela)
                _corr = dados_supabase.listar_correcoes(tabela)
                _nt = dados_supabase.listar_nao_trabalhados()
                _conf, _sug, _nao, _nao_trab = processar_hibrido(_ib, _cat, _ia, _corr, _nt)
                _regras = dados_supabase.listar_regras_st()
                _cf = _C.get("consumidor_final", False)
                for _it in _conf:
                    _it["st_unit"] = dados_supabase.calcular_st_difal(
                        _it["preco"], _it.get("ncm",""), uf_destino,
                        regras=_regras, consumidor_final=_cf)
                st.session_state["cotacao"].update({
                    "conf":_conf,"sug":_sug,"nao":_nao,"nao_trab":_nao_trab,
                    "subtotal":sum(i["total"] for i in _conf),
                    "total_st":sum(i.get("st_unit",0.0)*i["quantidade"] for i in _conf)})
                flash(f"{n_salvos} item(ns) corrigido(s)/confirmado(s) e aprendido(s). Cotação refeita.", "success")
                st.rerun()
            else:
                st.info("Nenhuma correção nova — os itens marcados como OK foram mantidos.")


# ── Rodapé ─────────────────────────────────────────────────────────────────
st.markdown("<br>", unsafe_allow_html=True)
st.markdown(f"""<p style='text-align:center;color:#aaa;font-size:11px;'>
  <span style='color:{COPPER};font-weight:600'>PASQUETTI</span>
  Comercio de Materiais Hidráulicos Ltda
  · tubos e conexões de ferro fundido ·
  Sistema de Cotações v2.0
</p>""", unsafe_allow_html=True)
